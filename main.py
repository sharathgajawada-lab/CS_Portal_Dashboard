"""
CS Portal Analytics — main.py
hear.com · Customer Support Intelligence

CALL BUDGET:
  Full refresh — fully sequential, 1s gap between each call
    - 9  time-series (KPIs, one per event type)
    - 3  top-n data   (articles, feedback, categories)
    - 3  top-n user IDs (video, search, article — to build user list)
    = 15 fixed calls
    + up to 10 content timelines  (Supabase enabled)
    + up to  2 feedback timelines (first 2 users only)
    = up to 27 calls total with Supabase / 21 without
  Schedule: every 2 hours
  Per page load: 0 calls — always served from cache
  On CMS error: serve stale cache, retry next cycle
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Header
from fastapi.responses import HTMLResponse, Response, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from collections import defaultdict, Counter
import httpx, os, asyncio, time, json, hashlib, csv, secrets, io, re

def _utcnow() -> datetime:
    """Naive UTC timestamp helper for legacy cache payload compatibility."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


# ── Supabase client (lazy init) ───────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
_sb_enabled  = bool(SUPABASE_URL and SUPABASE_KEY)

async def _sb_request(method: str, path: str, body: dict = None) -> dict:
    """Make a request to Supabase REST API."""
    if not _sb_enabled:
        return {}
    headers = {
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "return=minimal",
    }
    url = f"{SUPABASE_URL}/rest/v1{path}"
    async with httpx.AsyncClient(timeout=30) as client:
        if method == "GET":
            r = await client.get(url, headers=headers)
        elif method == "POST":
            headers["Prefer"] = "return=representation"
            r = await client.post(url, headers=headers, json=body)
        elif method == "UPSERT":
            headers["Prefer"] = "resolution=merge-duplicates,return=minimal"
            r = await client.post(url, headers=headers, json=body)
        else:
            return {}
        if r.status_code in (200, 201, 204):
            try: return r.json()
            except: return {}
        print(f"[supabase] {method} {path} → {r.status_code}: {r.text[:200]}", flush=True)
        return {}

async def _sb_ensure_tables():
    """Create tables if they don't exist using Supabase SQL API."""
    if not _sb_enabled:
        return
    sql = """
    CREATE TABLE IF NOT EXISTS cs_user_timelines (
        user_id      TEXT NOT NULL,
        project      TEXT NOT NULL,
        event_type   TEXT,
        item_id      TEXT,
        session_id   TEXT,
        ts           BIGINT,
        event_date   TEXT,
        properties   JSONB,
        fetched_at   TIMESTAMPTZ DEFAULT NOW(),
        PRIMARY KEY (user_id, project, ts)
    );
    CREATE TABLE IF NOT EXISTS cs_user_fetch_log (
        user_id      TEXT PRIMARY KEY,
        last_fetched TIMESTAMPTZ DEFAULT NOW(),
        event_count  INT DEFAULT 0
    );
    CREATE INDEX IF NOT EXISTS idx_timelines_date ON cs_user_timelines(event_date);
    CREATE INDEX IF NOT EXISTS idx_timelines_user ON cs_user_timelines(user_id);
    CREATE INDEX IF NOT EXISTS idx_timelines_type ON cs_user_timelines(event_type);
    """
    headers = {
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
    }
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.post(
            f"{SUPABASE_URL}/rest/v1/rpc/exec_sql",
            headers=headers,
            json={"sql": sql}
        )
        if r.status_code not in (200, 201, 204):
            # Try direct SQL via pg REST
            print(f"[supabase] table creation via RPC failed ({r.status_code}), trying direct", flush=True)
    print(f"[supabase] tables ready", flush=True)

async def _sb_get_unfetched_users(all_user_ids: list, limit: int = 10) -> list:
    """Return users we haven't fetched yet (or haven't fetched in 48h)."""
    if not _sb_enabled or not all_user_ids:
        return all_user_ids[:limit]
    # Get users we've already fetched recently
    result = await _sb_request("GET",
        "/cs_user_fetch_log?select=user_id,last_fetched&order=last_fetched.asc")
    if isinstance(result, list):
        fetched_recently = {
            r["user_id"] for r in result
            if r.get("last_fetched") and
            (_utcnow() - datetime.fromisoformat(
                r["last_fetched"].replace("Z","").split("+")[0]
            )).total_seconds() < 48 * 3600
        }
        # Prioritise users not yet fetched
        unfetched = [u for u in all_user_ids if u not in fetched_recently]
        if not unfetched:
            # All fetched recently — refresh oldest ones (preserve Supabase's asc order)
            known_set = set(all_user_ids)
            unfetched = [r["user_id"] for r in result if r["user_id"] in known_set]
        return unfetched[:limit]
    return all_user_ids[:limit]

async def _sb_store_events(user_id: str, project: str, events: list):
    """Store timeline events for a user in Supabase."""
    if not _sb_enabled or not events:
        return
    rows = []
    for ev in events:
        ts = ev.get("timestamp", 0)
        if not ts:
            continue
        from datetime import datetime as dt
        try:
            event_date = dt.utcfromtimestamp(ts / 1000).strftime("%Y-%m-%d")
        except Exception:
            continue
        rows.append({
            "user_id":    user_id,
            "project":    project,
            "event_type": _etype(ev),
            "item_id":    ev.get("item_id") or ev.get("itemId") or "",
            "session_id": ev.get("session_id") or ev.get("sessionId") or "",
            "ts":         int(ts),
            "event_date": event_date,
            "properties": _props(ev),
        })
    if not rows:
        return
    # Upsert in batches of 500
    for i in range(0, len(rows), 500):
        batch = rows[i:i+500]
        await _sb_request("UPSERT", "/cs_user_timelines", batch)
    # Update fetch log
    await _sb_request("UPSERT", "/cs_user_fetch_log", [{
        "user_id":    user_id,
        "last_fetched": _utcnow().isoformat() + "Z",
        "event_count": len(rows),
    }])
    print(f"[supabase] stored {len(rows)} events for {user_id[:16]}", flush=True)

async def _sb_get_all_events(date_from: str = "", date_to: str = "", project: str = "cs-portal-content-events") -> list:
    """Fetch stored timeline events from Supabase, optionally filtered by date/project."""
    if not _sb_enabled:
        return []
    path = "/cs_user_timelines?select=user_id,event_type,item_id,session_id,ts,event_date,properties"
    if project:
        path += f"&project=eq.{project}"
    if date_from:
        path += f"&event_date=gte.{date_from}"
    if date_to:
        path += f"&event_date=lte.{date_to}"
    path += "&limit=100000&order=ts.asc"
    result = await _sb_request("GET", path)
    if isinstance(result, list):
        print(f"[supabase] loaded {len(result)} events from DB", flush=True)
        return result
    return []

async def _sb_get_user_count() -> int:
    """Get total number of unique users stored."""
    if not _sb_enabled:
        return 0
    result = await _sb_request("GET", "/cs_user_fetch_log?select=user_id")
    return len(result) if isinstance(result, list) else 0



# ── Config ─────────────────────────────────────────────────────────────────────
CMS_BASE      = os.environ.get("CMS_BASE_URL", "https://cms.audibene.net/api/metrics")
API_KEY       = os.environ.get("CMS_API_KEY", "")
DATA_START    = os.environ.get("DATA_START", "2026-04-24")
APP_ENV       = os.environ.get("APP_ENV", "development").strip().lower()
DASHBOARD_VERSION = os.environ.get("DASHBOARD_VERSION", "2026.06-secure-metrics")

def _is_test_mode() -> bool:
    """Return True under pytest/local contract tests to avoid spawning network refresh tasks."""
    return (
        APP_ENV == "test"
        or API_KEY.startswith("TEST-")
        or bool(os.environ.get("PYTEST_CURRENT_TEST"))
    )

def _env_truthy(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "y", "on"}

# Sensitive/admin operations require a bearer or X-Admin-Token value.
# No insecure default token is provided. For production, set DASHBOARD_ADMIN_TOKEN.
# CSAT_UPLOAD_PASSWORD is accepted only as a backwards-compatible env alias.
ADMIN_TOKEN = (os.environ.get("DASHBOARD_ADMIN_TOKEN")
               or os.environ.get("CSAT_UPLOAD_PASSWORD")
               or "").strip()
ADMIN_TOKEN_REQUIRED = not _env_truthy("DISABLE_ADMIN_AUTH", False)
# Product decision for this dashboard: CSAT raw drilldowns are visible in the UI
# without an extra unlock step. Keep the global admin guard for debug/cache/full
# refresh endpoints, but do not block CSAT drilldown UX behind a modal.
CSAT_RAW_PUBLIC = _env_truthy("CSAT_RAW_PUBLIC", True)
CSAT_REFRESH_REQUIRES_ADMIN = _env_truthy("CSAT_REFRESH_REQUIRES_ADMIN", False)

async def require_admin_token(
    x_admin_token: str = Header(default=""),
    authorization: str = Header(default=""),
) -> bool:
    """Require an admin token for sensitive data and operational endpoints."""
    if not ADMIN_TOKEN_REQUIRED:
        return True
    if not ADMIN_TOKEN:
        raise HTTPException(status_code=503, detail="DASHBOARD_ADMIN_TOKEN is not configured")
    token = (x_admin_token or "").strip()
    if not token and authorization.lower().startswith("bearer "):
        token = authorization[7:].strip()
    if not token or not secrets.compare_digest(token, ADMIN_TOKEN):
        raise HTTPException(status_code=401, detail="Admin token required")
    return True

# ── Starter Guide Service ───────────────────────────────────────────────────────
SG_BASE       = os.environ.get("STARTER_GUIDE_BASE_URL", "https://starter-guide-service.audibene.net")
# JWT bearer token for authenticated /api/v1/* endpoints (set STARTER_GUIDE_API_TOKEN env var).
# Public /public/v1/* endpoints work without this token.
SG_API_TOKEN  = os.environ.get("STARTER_GUIDE_API_TOKEN", "")

# ── CSAT Survey Data ────────────────────────────────────────────────────────────
# Domo-first CSAT source. Startup may load bundled call_quality.csv as a fallback.
# POST /upload/csat is retained only as a protected break-glass ingestion path.
_csat_rows: list = []
_csat_index: dict = {}  # pre-built day-level index for O(days) not O(rows) queries

# Raw CSAT drilldown and Domo refresh are open by default for internal dashboard use;
# break-glass upload/debug/cache endpoints remain admin-token protected.

# Path where csat_index.json is saved for serving
CSAT_JSON_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "csat_index.json")

# Optional CS team allow-list. Future-proof default: do not drop newly-created
# teams unless CSAT_TEAMS is explicitly configured as a comma-separated allow-list.
# This prevents silent data loss when Operations adds a new team.
_DEFAULT_CSAT_TEAMS = ""
CSAT_TEAMS_ALLOW = [
    t.strip() for t in os.getenv("CSAT_TEAMS", _DEFAULT_CSAT_TEAMS).split(",") if t.strip()
]
# Voice AI is a first-class CS team. If an allow-list is configured in Render,
# keep it from being accidentally filtered out by older environment values.
if CSAT_TEAMS_ALLOW:
    _voice_ai_seen = any("".join(ch for ch in t.lower() if ch.isalnum()) in {"voiceai", "teamvoiceai"} for t in CSAT_TEAMS_ALLOW)
    if not _voice_ai_seen:
        CSAT_TEAMS_ALLOW.append("Voice AI")
_CSAT_TEAMS_ALLOW_LC = {t.lower() for t in CSAT_TEAMS_ALLOW}
CSAT_INCLUDE_BOT_CONSULTANTS = _env_truthy("CSAT_INCLUDE_BOT_CONSULTANTS", False)

CSAT_SCHEMA_VERSION = 2
CSAT_INDEX_KEY_LEGEND = {
    "t": "survey_count",
    "sr": "sum_rating",
    "s": "survey_solved_count",
    "l": "low_rating_count_1_or_2",
    "d": "rating_distribution",
    "tm": "teams",
    "cn": "consultants",
    "rs": "call_reasons",
    "uo": "unique_opportunities",
    "ft": "true_fcr_eligible_first_calls",
    "fr": "true_fcr_resolved_first_calls",
}
CSAT_METRIC_DEFINITIONS = {
    "csat.avg_rating": {
        "label": "Average CSAT",
        "grain": "survey_response",
        "formula": "sum(rating) / count(valid survey responses)",
        "scale": "1-5",
    },
    "csat.solved_rate": {
        "label": "Solved rate",
        "grain": "survey_response",
        "formula": "count(SOLVED=true) / count(valid survey responses)",
        "note": "This is not true first-call resolution.",
    },
    "csat.true_fcr": {
        "label": "True first-call resolution",
        "grain": "opportunity",
        "formula": "first call is solved and the opportunity has no later calls",
        "requires": ["OPPORTUNITY_ID"],
    },
}


def _canonical_csat_team(team: str) -> str:
    """Normalize common Domo/team-label variants without dropping new teams.

    This prevents Voice AI from disappearing when the source sends variants like
    VoiceAI, Voice-AI, or voice ai, while leaving unknown future teams intact.
    """
    raw = str(team or "").strip()
    if not raw:
        return ""
    compact = re.sub(r"[^a-z0-9]+", "", raw.lower())
    known = {
        "voiceai": "Voice AI",
        "teamvoiceai": "Voice AI",
        "hear4life": "Team Hear4Life",
        "teamhear4life": "Team Hear4Life",
        "amplifiers": "Team Amplifiers",
        "teamamplifiers": "Team Amplifiers",
        "soundcheck": "Team Sound Check",
        "teamsoundcheck": "Team Sound Check",
    }
    return known.get(compact, raw)


def _is_frank_ai_consultant(value: str) -> bool:
    """Return True for consultants that belong to Voice AI by name/ID.

    Product rule: every consultant whose name or ID says Frank AI is part of
    the Voice AI team, even if Domo sends a stale/blank/different team value.
    This catches variants like Frank Ai, Frank-AI, FrankAI, and Frank AI Bot.
    """
    compact = re.sub(r"[^a-z0-9]+", "", str(value or "").lower())
    return "frankai" in compact


def _normalize_csat_rows(rows: list) -> list:
    out = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        r = dict(row)
        team = _canonical_csat_team(r.get("team"))
        if _is_frank_ai_consultant(r.get("name")) or _is_frank_ai_consultant(r.get("cid")):
            team = "Voice AI"
        r["team"] = team
        out.append(r)
    return out


def _filter_allowed_teams(rows: list) -> list:
    """Keep only rows whose CONSULTANT_TEAM is in the allow-list (case-insensitive)."""
    if not _CSAT_TEAMS_ALLOW_LC:
        return rows
    kept = [r for r in rows if (r.get("team") or "").strip().lower() in _CSAT_TEAMS_ALLOW_LC]
    dropped = len(rows) - len(kept)
    if dropped:
        print(f"[csat] team filter: kept {len(kept)} rows, dropped {dropped} "
              f"(allow-list: {CSAT_TEAMS_ALLOW})", flush=True)
    return kept



def _build_csat_index(rows: list) -> dict:
    """Build the pre-aggregated day/week index from raw survey rows.
    Used by every data path — Domo pull, startup CSV, and manual upload.
    """
    global _csat_rows, _csat_index

    raw_rows = _normalize_csat_rows(list(rows or []))
    raw_row_count = len(raw_rows)

    # Drop teams/departments outside the CS allow-list only when CSAT_TEAMS is
    # explicitly configured. New teams otherwise flow through automatically.
    rows = _filter_allowed_teams(raw_rows)
    dropped_by_team_filter = raw_row_count - len(rows)

    rows = sorted(rows, key=lambda r: r.get("date", ""))
    _csat_rows = rows

    # ── True FCR pre-pass ────────────────────────────────────────────────────
    # "First-call resolved" = the FIRST call on an opportunity was SOLVED and the
    # opportunity has NO later call ever (any follow-up call = not resolved).
    # We tag each row: fcr_eval=True only for the FIRST call of each opportunity
    # (that first call is the unit we score), and fcr_ok=True if resolved.
    # Rows with no opp_id are not eligible (fcr_eval stays False) so true FCR
    # simply has no data until OPPORTUNITY_ID is present in the dataset.
    _opp_calls = {}
    for _r in rows:
        oid = (_r.get("opp_id") or "").strip()
        _r["_fcr_eval"] = False
        _r["_fcr_ok"] = False
        if oid:
            _opp_calls.setdefault(oid, []).append(_r)
    for oid, calls in _opp_calls.items():
        # earliest call by full datetime (fallback to date) is the "first contact"
        calls.sort(key=lambda x: (x.get("datetime") or x.get("date") or ""))
        first = calls[0]
        first["_fcr_eval"] = True
        # resolved only if the first call was solved AND it's the only call ever
        first["_fcr_ok"] = bool(first.get("solved")) and len(calls) == 1

    BAD_NAMES = {"none", "null", "", "n/a"}
    day_map   = {}
    week_cons = {}
    day_opp_sets = {}
    team_opp_sets = {}
    cons_opp_sets = {}

    import datetime as _dt_mod

    for r in rows:
        d    = r["date"]
        team = r["team"]
        cid  = r["cid"]
        name = r["name"]
        bad_team = not team or team.strip().lower() in BAD_NAMES
        name_lc = (name or "").strip().lower()
        is_frank_ai = _is_frank_ai_consultant(name) or _is_frank_ai_consultant(cid)
        if is_frank_ai and _canonical_csat_team(team) != "Voice AI":
            team = "Voice AI"
            r["team"] = "Voice AI"
        bad_name = (
            not name
            or name_lc in BAD_NAMES
        )

        if d not in day_map:
            day_map[d] = {"t":0,"sr":0,"s":0,"l":0,"d":{},"tm":{},"cn":{},"rs":{},"ft":0,"fr":0}
        dm = day_map[d]
        reason = (r.get("reason") or "").strip() or "Unspecified"
        summary = (r.get("summary") or r.get("summary_raw") or "").strip()
        call_id_for_reason = (r.get("call_id") or "").strip()
        opp_id = (r.get("opp_id") or "").strip()
        if opp_id:
            day_opp_sets.setdefault(d, set()).add(opp_id)
        dm["t"]  += 1
        dm["sr"] += r["rating"]
        dm["s"]  += int(r["solved"])
        dm["l"]  += int(r["rating"] <= 2)
        dm["d"][r["rating"]] = dm["d"].get(r["rating"], 0) + 1
        if reason not in dm["rs"]:
            dm["rs"][reason] = {"t":0,"sr":0,"s":0,"l":0,"d":{1:0,2:0,3:0,4:0,5:0},"sm":[]}
        dm["rs"][reason]["t"] += 1
        dm["rs"][reason]["sr"] += r["rating"]
        dm["rs"][reason]["s"] += int(r["solved"])
        dm["rs"][reason]["l"] += int(r["rating"] <= 2)
        dm["rs"][reason]["d"][r["rating"]] = dm["rs"][reason]["d"].get(r["rating"], 0) + 1
        # Store summary if either formatted or raw text exists
        display_summary = summary or r.get("summary_raw", "")[:1200]
        if display_summary and len(dm["rs"][reason]["sm"]) < 20:
            dm["rs"][reason]["sm"].append({
                "dt": r.get("datetime") or d,  # full datetime if available
                "n": (name or "").strip(),
                "cid": cid,  # consultant ID
                "team": team.strip(),  # consultant team
                "i": call_id_for_reason,
                "oid": opp_id,  # opportunity ID
                "rid": r.get("response_id", "").strip(),  # response ID
                "cby": r.get("created_by_id", "").strip(),  # created by ID
                "own": r.get("owner_id", "").strip(),  # owner ID
                "rat": r["rating"],  # rating
                "sol": int(r["solved"]),  # solved
                "t": display_summary,  # summary text (formatted or raw JSON)
                "summary": display_summary,  # compatibility alias for frontend consumers
                "raw": str(r.get("summary_raw") or "").strip(),  # preserve raw payload when present
            })
        # True FCR counters (only the first call of each opportunity is evaluated)
        if r.get("_fcr_eval"):
            dm["ft"] += 1
            dm["fr"] += int(r.get("_fcr_ok", False))

        if not bad_team:
            t = team.strip()
            if opp_id:
                team_opp_sets.setdefault((d, t), set()).add(opp_id)
            if t not in dm["tm"]:
                dm["tm"][t] = {"t":0,"sr":0,"s":0,"l":0,"d":{1:0,2:0,3:0,4:0,5:0},"rs":{},"ft":0,"fr":0}
            dm["tm"][t]["t"]  += 1
            dm["tm"][t]["sr"] += r["rating"]
            dm["tm"][t]["s"]  += int(r["solved"])
            dm["tm"][t]["l"]  += int(r["rating"] <= 2)
            dm["tm"][t]["d"][r["rating"]] = dm["tm"][t]["d"].get(r["rating"], 0) + 1
            if reason not in dm["tm"][t]["rs"]:
                dm["tm"][t]["rs"][reason] = {"t":0,"sr":0,"s":0,"l":0,"d":{1:0,2:0,3:0,4:0,5:0}}
            dm["tm"][t]["rs"][reason]["t"] += 1
            dm["tm"][t]["rs"][reason]["sr"] += r["rating"]
            dm["tm"][t]["rs"][reason]["s"] += int(r["solved"])
            dm["tm"][t]["rs"][reason]["l"] += int(r["rating"] <= 2)
            dm["tm"][t]["rs"][reason]["d"][r["rating"]] = dm["tm"][t]["rs"][reason]["d"].get(r["rating"], 0) + 1
            if r.get("_fcr_eval"):
                dm["tm"][t]["ft"] += 1
                dm["tm"][t]["fr"] += int(r.get("_fcr_ok", False))

        if not bad_name:
            # Per-day consultant data — ensures team totals == sum of consultant totals
            if opp_id:
                cons_opp_sets.setdefault((d, cid), set()).add(opp_id)
            if cid not in dm["cn"]:
                dm["cn"][cid] = {"n":name,"tm":team.strip(),"t":0,"sr":0,"s":0,"l":0,"d":{1:0,2:0,3:0,4:0,5:0},"rs":{},"c":[],"ft":0,"fr":0}
            dm["cn"][cid]["t"]  += 1
            dm["cn"][cid]["sr"] += r["rating"]
            dm["cn"][cid]["s"]  += int(r["solved"])
            dm["cn"][cid]["l"]  += int(r["rating"] <= 2)
            dm["cn"][cid]["d"][r["rating"]] = dm["cn"][cid]["d"].get(r["rating"], 0) + 1
            if reason not in dm["cn"][cid]["rs"]:
                dm["cn"][cid]["rs"][reason] = {"t":0,"sr":0,"s":0,"l":0,"d":{1:0,2:0,3:0,4:0,5:0}}
            dm["cn"][cid]["rs"][reason]["t"] += 1
            dm["cn"][cid]["rs"][reason]["sr"] += r["rating"]
            dm["cn"][cid]["rs"][reason]["s"] += int(r["solved"])
            dm["cn"][cid]["rs"][reason]["l"] += int(r["rating"] <= 2)
            dm["cn"][cid]["rs"][reason]["d"][r["rating"]] = dm["cn"][cid]["rs"][reason]["d"].get(r["rating"], 0) + 1
            if r.get("_fcr_eval"):
                dm["cn"][cid]["ft"] += 1
                dm["cn"][cid]["fr"] += int(r.get("_fcr_ok", False))
            # Per-call record for individual-call tables (Call ID → UCJ link).
            # Compact keys: i=call_id, r=rating, s=solved(0/1), rs=reason. Date comes from day key.
            call_id = r.get("call_id") or ""
            if call_id:
                call_record = {"i": call_id, "r": r["rating"], "s": int(r["solved"])}
                if reason and reason != "Unspecified":
                    call_record["rs"] = reason
                if r.get("datetime") and r.get("datetime") != d:
                    call_record["dt"] = r.get("datetime")
                if display_summary:
                    call_record["t"] = display_summary
                dm["cn"][cid]["c"].append(call_record)

        if not bad_name:
            try:
                parts = d.split("-")
                dt  = _dt_mod.date(int(parts[0]), int(parts[1]), int(parts[2]))
                wk  = (dt - _dt_mod.timedelta(days=dt.weekday())).isoformat()
            except Exception:
                continue
            if wk not in week_cons:
                week_cons[wk] = {}
            if cid not in week_cons[wk]:
                week_cons[wk][cid] = {"n":name,"tm":team,"t":0,"sr":0,"s":0,"l":0,"d":{1:0,2:0,3:0,4:0,5:0}}
            week_cons[wk][cid]["t"]  += 1
            week_cons[wk][cid]["sr"] += r["rating"]
            week_cons[wk][cid]["s"]  += int(r["solved"])
            week_cons[wk][cid]["l"]  += int(r["rating"] <= 2)
            week_cons[wk][cid]["d"][r["rating"]] = week_cons[wk][cid]["d"].get(r["rating"], 0) + 1

    # Unique opportunities at day / team / consultant level.
    for d, dm in day_map.items():
        dm["uo"] = len(day_opp_sets.get(d, set()))
        for t, tv in (dm.get("tm") or {}).items():
            tv["uo"] = len(team_opp_sets.get((d, t), set()))
        for cid, cv in (dm.get("cn") or {}).items():
            cv["uo"] = len(cons_opp_sets.get((d, cid), set()))

    dates = sorted(day_map.keys())
    quality = {
        "raw_rows": raw_row_count,
        "indexed_rows": len(rows),
        "dropped_by_team_filter": dropped_by_team_filter,
        "missing_team_rows": sum(1 for r in rows if not (r.get("team") or "").strip()),
        "missing_consultant_name_rows": sum(1 for r in rows if not (r.get("name") or "").strip()),
        "missing_call_id_rows": sum(1 for r in rows if not (r.get("call_id") or "").strip()),
        "missing_opportunity_id_rows": sum(1 for r in rows if not (r.get("opp_id") or "").strip()),
        "missing_summary_rows": sum(1 for r in rows if not ((r.get("summary") or "").strip() or (r.get("summary_raw") or "").strip())),
        "distinct_teams": len({(r.get("team") or "").strip() for r in rows if (r.get("team") or "").strip()}),
        "distinct_consultants": len({(r.get("cid") or "").strip() for r in rows if (r.get("cid") or "").strip()}),
        "distinct_days": len(day_map),
        "true_fcr_available": any((dm.get("ft") or 0) > 0 for dm in day_map.values()),
        "team_filter": {
            "enabled": bool(CSAT_TEAMS_ALLOW),
            "allow_list": CSAT_TEAMS_ALLOW,
        },
        "bot_consultants_included": CSAT_INCLUDE_BOT_CONSULTANTS,
        "frank_ai_consultants_mapped_to_voice_ai": True,
    }
    index_data = {
        "available":  True,
        "schema_version": CSAT_SCHEMA_VERSION,
        "key_legend": CSAT_INDEX_KEY_LEGEND,
        "metric_definitions": CSAT_METRIC_DEFINITIONS,
        "quality": quality,
        "date_min":   dates[0]  if dates else "",
        "date_max":   dates[-1] if dates else "",
        "days":       day_map,
        "week_cons":  week_cons,
        "total_rows": len(rows),
        "raw_total_rows": raw_row_count,
        "generated":  _utcnow().isoformat() + "Z",
    }

    # Also keep the old _csat_index format for the /api/csat endpoint
    _csat_index["day_map"]   = day_map
    _csat_index["date_min"]  = index_data["date_min"]
    _csat_index["date_max"]  = index_data["date_max"]
    _csat_index["dates"]     = [r["date"] for r in rows]
    _csat_index["rows"]      = rows
    _csat_index["index_data"] = index_data   # cached for /api/csat/raw

    print(f"[csat] index built: {len(rows):,} rows · {len(day_map)} days · {len(week_cons)} weeks", flush=True)
    return index_data

_last_csv_columns = []  # Global to track columns for debugging
_last_raw_sample = {}   # Global to store first raw row for inspection

def _parse_csv_text(text: str) -> list:
    """Parse CSAT CSV text (from disk or Domo export) into survey rows."""
    global _last_csv_columns
    def _canon_key(value) -> str:
        return "".join(ch for ch in str(value or "") if ch.isalnum()).upper()

    def _normalize_summary(raw_value) -> str:
        raw = str(raw_value or "").strip()
        if not raw:
            return ""
        # FULL_SUMMARY_JSON__C may be JSON or plain text; prefer a readable text field.
        if raw[:1] in ("{", "["):
            try:
                payload = json.loads(raw)
                # Some exports wrap JSON as an encoded string (JSON-in-JSON).
                if isinstance(payload, str):
                    p2 = payload.strip()
                    if p2[:1] in ("{", "["):
                        payload = json.loads(p2)
                if isinstance(payload, dict):
                    # Common AI summary payloads can vary by key casing.
                    p = {str(k).strip().lower(): v for k, v in payload.items()}

                    def _as_text(v):
                        if v is None:
                            return ""
                        if isinstance(v, str):
                            return v.strip()
                        if isinstance(v, (int, float, bool)):
                            return str(v)
                        if isinstance(v, (dict, list)):
                            return json.dumps(v, ensure_ascii=False)
                        return str(v).strip()

                    def getv(*ks):
                        for k in ks:
                            val = _as_text(p.get(k.lower()))
                            if val:
                                return val
                        return ""

                    call_summary = getv("callsummary", "call_summary", "summary", "full_summary", "conversation_summary")
                    call_reason  = getv("customerscallreason", "customer_call_reason", "callreason", "reason")
                    next_steps   = getv("nextsteps", "next_steps", "actionitems", "actions")

                    if call_summary or call_reason or next_steps:
                        parts = []
                        if call_reason:
                            parts.append("Customer Call Reason:\n" + call_reason)
                        if call_summary:
                            parts.append("Call Summary:\n" + call_summary)
                        if next_steps:
                            parts.append("Next Steps:\n" + next_steps)
                        return "\n\n".join(parts)[:2200]

                keys = {
                    "summary", "call_summary", "callsummary", "full_summary", "short_summary",
                    "conversation_summary", "customerscallreason", "nextsteps",
                    "text", "content", "overview",
                }
                def _walk(obj):
                    if isinstance(obj, str):
                        s = obj.strip()
                        return s if len(s) > 8 else ""
                    if isinstance(obj, dict):
                        for k, v in obj.items():
                            if str(k).strip().lower() in keys:
                                got = _walk(v)
                                if got:
                                    return got
                        for v in obj.values():
                            got = _walk(v)
                            if got:
                                return got
                    if isinstance(obj, list):
                        for v in obj:
                            got = _walk(v)
                            if got:
                                return got
                    return ""
                extracted = _walk(payload)
                if extracted:
                    return extracted[:2200]
            except Exception:
                pass
        return raw[:2200]

    rows = []
    first_row = True
    for row in csv.DictReader(io.StringIO(text)):
        if first_row:
            global _last_raw_sample
            _last_csv_columns = list(row.keys())
            _last_raw_sample = dict(row)  # Store first raw row for inspection
            print(f"[parse_csv] columns detected ({len(_last_csv_columns)}): {_last_csv_columns}", flush=True)
            print(f"[parse_csv] first row raw data: {dict(list(row.items())[:7])}", flush=True)
            first_row = False
        try:
            # Canonicalize headers to tolerate spaces, punctuation, and case changes.
            row_u = {str(k or "").strip().upper(): v for k, v in row.items()}
            row_canon = {_canon_key(k): v for k, v in row.items()}

            def _gv(*keys):
                for k in keys:
                    v = row_u.get(str(k).strip().upper())
                    if v is None or str(v).strip() == "":
                        v = row_canon.get(_canon_key(k))
                    if v is not None and str(v).strip() != "":
                        return v
                return ""

            raw_date = _gv("DATE", "DATETIME") or ""
            date_str = raw_date[:10]  # take just YYYY-MM-DD from datetime
            # opportunity id (for true FCR) — tolerate a few likely column names.
            opp = _gv("OPPORTUNITY_ID", "OPPORTUNITYID", "OPPORTUNITY")
            summary_raw = _gv("FULL_SUMMARY_JSON__C", "FULL_SUMMARY_JSON", "CALL_SUMMARY", "CALLSUMMARY", "SUMMARY")
            if not str(summary_raw or "").strip():
                # Last-resort scan in case Domo renamed summary headers unexpectedly.
                summary_candidates = []
                for k, v in row_canon.items():
                    if v is None or str(v).strip() == "":
                        continue
                    if any(tok in k for tok in ("FULLSUMMARY", "CALLSUMMARY", "SUMMARYJSON", "SUMMARYTEXT", "TRANSCRIPT", "NOTES")):
                        summary_candidates.append(str(v).strip())
                if summary_candidates:
                    summary_raw = max(summary_candidates, key=len)
            if not str(summary_raw or "").strip():
                # Heuristic fallback: pick the longest narrative-looking text cell.
                skip_tokens = {
                    "RATING", "CONSULTANTID", "CONSULTANTNAME", "CONSULTANTTEAM", "DATE", "DATETIME",
                    "SOLVED", "CALLID", "OPPORTUNITYID", "OWNERID", "CREATEDBYID", "RESPONSEID", "REASONFORCALL",
                }
                for k, v in row_canon.items():
                    if not v:
                        continue
                    txt = str(v).strip()
                    if len(txt) < 80:
                        continue
                    if _canon_key(k) in skip_tokens:
                        continue
                    # Prefer JSON-ish payloads and multi-sentence text.
                    if txt[:1] in ("{", "[") or txt.count(" ") >= 12:
                        summary_raw = txt
                        break
            if not str(summary_raw or "").strip():
                call_reason_raw = str(_gv("CUSTOMERSCALLREASON", "CUSTOMER_CALL_REASON", "CALLREASON", "REASON_FOR_CALL__C", "REASON_FOR_CALL", "CALL_REASON", "REASON")).strip()
                call_summary_raw = str(_gv("CALLSUMMARY", "CALL_SUMMARY", "SUMMARY", "FULL_SUMMARY")).strip()
                next_steps_raw = str(_gv("NEXTSTEPS", "NEXT_STEPS", "ACTIONITEMS", "ACTIONS")).strip()
                if call_reason_raw or call_summary_raw or next_steps_raw:
                    parts = []
                    if call_reason_raw:
                        parts.append("Customer Call Reason:\n" + call_reason_raw)
                    if call_summary_raw:
                        parts.append("Call Summary:\n" + call_summary_raw)
                    if next_steps_raw:
                        parts.append("Next Steps:\n" + next_steps_raw)
                    summary_raw = "\n\n".join(parts)
            rows.append({
                "rating": int(float(_gv("RATING"))),  # handles "5.0" and "5"
                "cid":    str(_gv("CONSULTANT_ID")).strip(),
                "name":   str(_gv("CONSULTANT_NAME")).strip(),
                "team":   _canonical_csat_team(_gv("CONSULTANT_TEAM")),
                "date":   date_str,
                "datetime": str(raw_date).strip(),  # full timestamp for FCR ordering
                "solved": str(_gv("SOLVED")).strip().lower() == "true",
                "call_id": str(_gv("CALL_ID")).strip(),
                "opp_id": str(opp).strip(),
                "reason": str(_gv("REASON_FOR_CALL__C", "REASON_FOR_CALL", "CALL_REASON", "REASON")).strip(),
                "summary": _normalize_summary(summary_raw),
                "summary_raw": str(summary_raw or "").strip(),
                "owner_id": str(_gv("OWNER_ID")).strip(),
                "created_by_id": str(_gv("CREATEDBYID", "CREATED_BY_ID")).strip(),
                "response_id": str(_gv("RESPONSE_ID")).strip(),
            })
        except (ValueError, KeyError):
            continue
    return rows


# ── Domo automated pull ────────────────────────────────────────────────────────
# Pulls the CSAT dataset straight from Domo on the 2-hour refresh cycle, so the
# manual "Update CSAT" upload is no longer required. Credentials come from env
# vars set in Render (never hard-coded):
#   DOMO_CLIENT_ID, DOMO_CLIENT_SECRET, DOMO_DATASET_ID
# The OAuth client must be created at developer.domo.com with the `data` scope
# (requires a Domo admin). Tokens last ~1 hour; we fetch a fresh one each pull.
DOMO_CLIENT_ID     = os.getenv("DOMO_CLIENT_ID", "")
DOMO_CLIENT_SECRET = os.getenv("DOMO_CLIENT_SECRET", "")
DOMO_DATASET_ID    = os.getenv("DOMO_DATASET_ID", "")
DOMO_API_BASE      = os.getenv("DOMO_API_BASE", "https://api.domo.com")
CSAT_REFRESH_SOURCE = {
    "source": "not_loaded",
    "loaded_at": "",
    "last_success_at": "",
    "last_error": "",
    "domo_configured": False,
}


def _mark_csat_source(source: str, error: str = "") -> None:
    """Track where the active CSAT index came from for dashboard source/status UX."""
    ts = _utcnow().isoformat() + "Z"
    CSAT_REFRESH_SOURCE.update({
        "source": source,
        "loaded_at": ts,
        "domo_configured": _domo_configured(),
        "last_error": error or "",
    })
    if not error:
        CSAT_REFRESH_SOURCE["last_success_at"] = ts


def _domo_configured() -> bool:
    return bool(DOMO_CLIENT_ID and DOMO_CLIENT_SECRET and DOMO_DATASET_ID)


def _domo_fetch_csv() -> str | None:
    """Authenticate to Domo and export the CSAT dataset as CSV text.

    Returns the CSV string on success, or None on any failure (caller falls
    back to the bundled CSV). Synchronous httpx — called from the refresh loop
    via asyncio.to_thread so it never blocks the event loop.
    """
    if not _domo_configured():
        return None
    try:
        with httpx.Client(timeout=60.0) as client:
            # 1) OAuth client-credentials grant → bearer token (Basic auth: id:secret).
            #    A 400 invalid_request usually means the requested scope doesn't match
            #    what the client was granted. So try a few strategies in order:
            #    (a) no scope → Domo returns the client's own granted scopes,
            #    (b) scope=data, (c) scope="data user". First 200 wins.
            token = None
            last_err = ""
            for scope in (None, "data", "data user"):
                params = {"grant_type": "client_credentials"}
                if scope:
                    params["scope"] = scope
                tok = client.get(
                    f"{DOMO_API_BASE}/oauth/token",
                    params=params,
                    auth=(DOMO_CLIENT_ID, DOMO_CLIENT_SECRET),
                )
                if tok.status_code == 200:
                    token = tok.json().get("access_token")
                    granted = tok.json().get("scope", "?")
                    print(f"[domo] auth ok (scope requested={scope or 'none'}, granted={granted})", flush=True)
                    break
                last_err = f"HTTP {tok.status_code} {tok.text[:150]}"
                print(f"[domo] auth attempt scope={scope or 'none'} failed: {last_err}", flush=True)
            if not token:
                print(f"[domo] auth failed on all scope attempts — last: {last_err}", flush=True)
                return None

            # 2) Export dataset as CSV (includeHeader so DictReader gets column names).
            exp = client.get(
                f"{DOMO_API_BASE}/v1/datasets/{DOMO_DATASET_ID}/data",
                params={"includeHeader": "true"},
                headers={"Authorization": f"Bearer {token}", "Accept": "text/csv"},
            )
            if exp.status_code != 200:
                print(f"[domo] export failed: HTTP {exp.status_code} {exp.text[:200]}", flush=True)
                return None
            csv_text = exp.text
            if not csv_text or "CONSULTANT_ID" not in csv_text:
                print("[domo] export returned unexpected/empty data — keeping previous", flush=True)
                return None
            print(f"[domo] exported dataset {DOMO_DATASET_ID[:8]}… ({len(csv_text)//1024} KB)", flush=True)
            return csv_text
    except Exception as e:
        print(f"[domo] pull error: {e}", flush=True)
        return None


def _refresh_csat_from_domo() -> bool:
    """Pull from Domo and rebuild the CSAT index. Returns True if it succeeded."""
    csv_text = _domo_fetch_csv()
    if csv_text is None:
        _mark_csat_source(CSAT_REFRESH_SOURCE.get("source") or "domo", "Domo export failed or returned no CSV")
        return False
    rows = _parse_csv_text(csv_text)
    if not rows:
        print("[domo] parsed 0 rows — keeping previous index", flush=True)
        _mark_csat_source(CSAT_REFRESH_SOURCE.get("source") or "domo", "Domo export parsed 0 valid rows")
        return False
    _build_csat_index(rows)
    _mark_csat_source("domo")
    _save_csat_json()
    print(f"[domo] CSAT index rebuilt from Domo — {len(rows)} rows", flush=True)
    return True


def _load_csat_csv():
    """Load CSAT data at startup. Prefer a live Domo pull; fall back to bundled CSV."""
    if _domo_configured():
        print("[csat] Domo configured — pulling at startup", flush=True)
        if _refresh_csat_from_domo():
            return
        print("[csat] Domo pull failed at startup — falling back to bundled CSV", flush=True)
    else:
        print("[csat] Domo not configured (set DOMO_* env vars) — using bundled CSV", flush=True)

    paths = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "call_quality.csv"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "call_quality.csv"),
        os.path.join(os.getcwd(), "call_quality.csv"),
        os.path.join(os.getcwd(), "data", "call_quality.csv"),
        "/app/call_quality.csv",
        "/app/data/call_quality.csv",
        "call_quality.csv",
        "data/call_quality.csv",
    ]
    for path in paths:
        if not os.path.exists(path):
            continue
        with open(path, newline='', encoding='utf-8') as f:
            rows = _parse_csv_text(f.read())
        _build_csat_index(rows)
        _mark_csat_source("bundled_csv")
        _save_csat_json()
        print(f"[csat] loaded from {path}", flush=True)
        return
    _mark_csat_source("not_loaded", "No bundled CSV found and Domo did not load")
    print("[csat] call_quality.csv not found — no CSAT index available", flush=True)


def _parse_excel_bytes(data: bytes) -> list:
    """Parse Excel file bytes into survey rows. Returns list of row dicts."""
    import datetime as _dt_mod

    def _canon_key(value) -> str:
        return "".join(ch for ch in str(value or "") if ch.isalnum()).upper()

    def _normalize_summary(raw_value) -> str:
        raw = str(raw_value or "").strip()
        if not raw:
            return ""
        if raw[:1] in ("{", "["):
            try:
                payload = json.loads(raw)
                # Some exports wrap JSON as an encoded string (JSON-in-JSON).
                if isinstance(payload, str):
                    p2 = payload.strip()
                    if p2[:1] in ("{", "["):
                        payload = json.loads(p2)
                if isinstance(payload, dict):
                    p = {str(k).strip().lower(): v for k, v in payload.items()}

                    def _as_text(v):
                        if v is None:
                            return ""
                        if isinstance(v, str):
                            return v.strip()
                        if isinstance(v, (int, float, bool)):
                            return str(v)
                        if isinstance(v, (dict, list)):
                            return json.dumps(v, ensure_ascii=False)
                        return str(v).strip()

                    def getv(*ks):
                        for k in ks:
                            val = _as_text(p.get(k.lower()))
                            if val:
                                return val
                        return ""

                    call_summary = getv("callsummary", "call_summary", "summary", "full_summary", "conversation_summary")
                    call_reason  = getv("customerscallreason", "customer_call_reason", "callreason", "reason")
                    next_steps   = getv("nextsteps", "next_steps", "actionitems", "actions")
                    if call_summary or call_reason or next_steps:
                        parts = []
                        if call_reason:
                            parts.append("Customer Call Reason:\n" + call_reason)
                        if call_summary:
                            parts.append("Call Summary:\n" + call_summary)
                        if next_steps:
                            parts.append("Next Steps:\n" + next_steps)
                        return "\n\n".join(parts)[:2200]
            except Exception:
                pass
        return raw[:2200]

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed — add to requirements.txt")

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not raw_rows:
        raise ValueError("Empty workbook")

    headers_raw = [str(h).strip() if h else "" for h in raw_rows[0]]
    headers = [h.upper() for h in headers_raw]
    col     = {name: i for i, name in enumerate(headers)}
    col_canon = {_canon_key(name): i for i, name in enumerate(headers_raw)}

    def _cell_value(row_data, *keys):
        for k in keys:
            idx = col.get(str(k).strip().upper())
            if idx is None:
                idx = col_canon.get(_canon_key(k))
            if idx is None or idx >= len(row_data):
                continue
            val = row_data[idx]
            if val is not None and str(val).strip() != "":
                return val
        return ""

    required = {"RATING", "CONSULTANT_ID", "CONSULTANT_NAME", "CONSULTANT_TEAM", "DATETIME", "SOLVED"}
    required_ok = set(col.keys()) | set(col_canon.keys())
    missing  = {k for k in required if _canon_key(k) not in required_ok and k not in required_ok}
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    rows = []
    for row in raw_rows[1:]:
        try:
            rating  = int(float(_cell_value(row, "RATING") or 0))
            cid     = str(_cell_value(row, "CONSULTANT_ID") or "").strip()
            name    = str(_cell_value(row, "CONSULTANT_NAME") or "").strip()
            team    = _canonical_csat_team(_cell_value(row, "CONSULTANT_TEAM") or "")
            dt_val  = _cell_value(row, "DATETIME", "DATE")
            solved  = str(_cell_value(row, "SOLVED") or "").strip().lower() == "true"
            call_id = str(_cell_value(row, "CALL_ID") or "").strip()
            opp_id  = str(_cell_value(row, "OPPORTUNITY_ID", "OPPORTUNITYID", "OPPORTUNITY") or "").strip()
            reason = str(_cell_value(row, "REASON_FOR_CALL__C", "REASON_FOR_CALL", "CALL_REASON", "REASON") or "").strip()
            owner_id = str(_cell_value(row, "OWNER_ID") or "").strip()
            created_by_id = str(_cell_value(row, "CREATEDBYID", "CREATED_BY_ID") or "").strip()
            response_id = str(_cell_value(row, "RESPONSE_ID") or "").strip()
            summary_raw = str(_cell_value(row, "FULL_SUMMARY_JSON__C", "FULL_SUMMARY_JSON", "CALL_SUMMARY", "CALLSUMMARY", "SUMMARY") or "").strip()

            if not summary_raw:
                summary_candidates = []
                for header_name, idx in col_canon.items():
                    if idx >= len(row):
                        continue
                    val = row[idx]
                    if val is None or str(val).strip() == "":
                        continue
                    if any(tok in header_name for tok in ("FULLSUMMARY", "CALLSUMMARY", "SUMMARYJSON", "SUMMARYTEXT", "TRANSCRIPT", "NOTES")):
                        summary_candidates.append(str(val).strip())
                if summary_candidates:
                    summary_raw = max(summary_candidates, key=len)

            if not summary_raw:
                skip_tokens = {
                    "RATING", "CONSULTANTID", "CONSULTANTNAME", "CONSULTANTTEAM", "DATE", "DATETIME",
                    "SOLVED", "CALLID", "OPPORTUNITYID", "OWNERID", "CREATEDBYID", "RESPONSEID", "REASONFORCALL",
                }
                for header_name, idx in col_canon.items():
                    if idx >= len(row):
                        continue
                    val = row[idx]
                    if val is None:
                        continue
                    txt = str(val).strip()
                    if len(txt) < 80:
                        continue
                    if header_name in skip_tokens:
                        continue
                    if txt[:1] in ("{", "[") or txt.count(" ") >= 12:
                        summary_raw = txt
                        break

            if not summary_raw:
                call_reason_raw = ""
                for _k in ("CUSTOMERSCALLREASON", "CUSTOMER_CALL_REASON", "CALLREASON", "REASON_FOR_CALL__C", "REASON_FOR_CALL", "CALL_REASON", "REASON"):
                    call_reason_raw = str(_cell_value(row, _k) or "").strip()
                    if call_reason_raw:
                        break
                call_summary_raw = ""
                for _k in ("CALLSUMMARY", "CALL_SUMMARY", "SUMMARY", "FULL_SUMMARY"):
                    call_summary_raw = str(_cell_value(row, _k) or "").strip()
                    if call_summary_raw:
                        break
                next_steps_raw = ""
                for _k in ("NEXTSTEPS", "NEXT_STEPS", "ACTIONITEMS", "ACTIONS"):
                    next_steps_raw = str(_cell_value(row, _k) or "").strip()
                    if next_steps_raw:
                        break
                if call_reason_raw or call_summary_raw or next_steps_raw:
                    parts = []
                    if call_reason_raw:
                        parts.append("Customer Call Reason:\n" + call_reason_raw)
                    if call_summary_raw:
                        parts.append("Call Summary:\n" + call_summary_raw)
                    if next_steps_raw:
                        parts.append("Next Steps:\n" + next_steps_raw)
                    summary_raw = "\n\n".join(parts)
            if isinstance(dt_val, (_dt_mod.datetime, _dt_mod.date)):
                date_str = dt_val.strftime("%Y-%m-%d")
                dt_full  = dt_val.isoformat()
            elif isinstance(dt_val, (int, float)):
                _d = (_dt_mod.date(1899, 12, 30) + _dt_mod.timedelta(days=float(dt_val)))
                date_str = _d.strftime("%Y-%m-%d")
                dt_full  = _d.isoformat()
            else:
                date_str = str(dt_val)[:10]
                dt_full  = str(dt_val).strip()
            if not (1 <= rating <= 5) or not cid or not date_str:
                continue
            rows.append({"rating":rating,"cid":cid,"name":name,
                         "team":team,"date":date_str,"datetime":dt_full,"solved":solved,
                         "call_id":call_id,"opp_id":opp_id,"reason":reason,
                         "summary":_normalize_summary(summary_raw),
                         "summary_raw":str(summary_raw or "").strip(),
                         "owner_id":owner_id,"created_by_id":created_by_id,"response_id":response_id})
        except (TypeError, ValueError):
            continue
    return rows


def _save_csat_json():
    """Save the current index to data/csat_index.json for serving."""
    index_data = _csat_index.get("index_data")
    if not index_data:
        return
    try:
        os.makedirs(os.path.dirname(CSAT_JSON_PATH), exist_ok=True)
        with open(CSAT_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(index_data, f, separators=(",",":"))
        print(f"[csat] saved index to {CSAT_JSON_PATH}", flush=True)
    except Exception as e:
        print(f"[csat] failed to save JSON: {e}", flush=True)



CACHE_TTL     = 7200   # 2 hr fresh
STALE_TTL     = 86400  # 24 hr stale — serve old data if CMS is down
REFRESH_SEC   = 7200   # refresh every 2 hours
CALL_GAP      = 1.0    # 1 second between every CMS call
SEARCH_GAP_MS = 3000

EVENTS = [
    {"key": "auth.login",             "project": "cs-portal-auth-events",       "label": "Logins",          "color": "#0e6e45"},
    {"key": "article.viewed",         "project": "cs-portal-content-events",    "label": "Articles viewed", "color": "#c47a0a"},
    {"key": "search.performed",       "project": "cs-portal-content-events",    "label": "Searches",        "color": "#5b3fbf"},
    {"key": "video.watched",          "project": "cs-portal-content-events",    "label": "Videos watched",  "color": "#c03030"},
    {"key": "category.viewed",        "project": "cs-portal-content-events",    "label": "Category views",  "color": "#1a4fdb"},
    {"key": "article.feedback",       "project": "cs-portal-feedback-events",   "label": "Feedback",        "color": "#0f6e56"},
    {"key": "order_supplies.visited", "project": "cs-portal-items-events",      "label": "Supplies visits", "color": "#b45309"},
    {"key": "auth.logout",            "project": "cs-portal-auth-events",       "label": "Logouts",         "color": "#6b7280"},
    {"key": "scheduling.started",     "project": "cs-portal-scheduling-events", "label": "Scheduling",      "color": "#7c3aed"},
]

# ── Cache ──────────────────────────────────────────────────────────────────────
_cache = {}
CACHE_FILE = "/tmp/cs_portal_cache.json"

def cache_get(key):
    e = _cache.get(key)
    if not e: return None, False
    age = time.time() - e["ts"]
    if age < CACHE_TTL:  return e["data"], True
    if age < STALE_TTL:  return e["data"], False
    return None, False

def cache_set(key, data):
    _cache[key] = {"ts": time.time(), "data": data}
    # Persist to disk so cache survives server restarts
    try:
        with open(CACHE_FILE, "w") as f:
            json.dump(_cache, f)
    except Exception as ex:
        print(f"[cache] disk write failed: {ex}", flush=True)

def load_cache_from_disk():
    """Load cache from disk on startup — survives Render restarts."""
    try:
        with open(CACHE_FILE) as f:
            data = json.load(f)
        _cache.update(data)
        keys = list(data.keys())
        print(f"[cache] loaded from disk: {keys}", flush=True)
    except FileNotFoundError:
        print("[cache] no disk cache found — fresh start", flush=True)
    except Exception as ex:
        print(f"[cache] disk read failed: {ex}", flush=True)

# ── HTTP client ────────────────────────────────────────────────────────────────
_client = None

def get_client():
    global _client
    if _client is None or _client.is_closed:
        _client = httpx.AsyncClient(
            timeout=20,
            limits=httpx.Limits(max_connections=3, max_keepalive_connections=2),
            headers={
                "api-key": API_KEY,
                "x-api-key": API_KEY,
                "Accept": "application/json",
            },
        )
    return _client

# ── Single lock — only ONE refresh at a time ───────────────────────────────────
_refresh_lock = None

def get_lock():
    global _refresh_lock
    if _refresh_lock is None:
        _refresh_lock = asyncio.Lock()
    return _refresh_lock

# ── Core HTTP — one call at a time ─────────────────────────────────────────────
async def _get(url, params, retries=3):
    client = get_client()
    for attempt in range(retries):
        try:
            r = await client.get(url, params=params)
            if r.status_code in (429, 502, 503, 504):
                wait = 5 * (attempt + 1)
                print(f"[CMS] {r.status_code} — waiting {wait}s", flush=True)
                await asyncio.sleep(wait)
                continue
            if r.status_code != 200:
                body_preview = (r.text or "").strip().replace("\n", " ")[:180]
                print(f"[CMS] {r.status_code} {url} :: {body_preview}", flush=True)
                return None
            text = r.text.strip()
            return json.loads(text) if text else None
        except Exception as ex:
            print(f"[CMS] error: {ex}", flush=True)
            await asyncio.sleep(5 * (attempt + 1))
    return None

async def _call(url, params):
    """Make one CMS call then wait CALL_GAP seconds before returning."""
    result = await _get(url, params)
    await asyncio.sleep(CALL_GAP)
    return result

async def _timeseries(project, event):
    data = await _call(f"{CMS_BASE}/{project}/query/time-series",
                       {"event": event, "bucket": "day"})
    if not data: return []
    daily = defaultdict(int)
    for p in data.get("series", []):
        ts = p.get("ts") or p.get("timestamp")
        if ts:
            try:
                d = datetime.utcfromtimestamp(int(ts)/1000).strftime("%Y-%m-%d")
                daily[d] += int(p.get("count", 0) or 0)
            except Exception:
                pass
    if not daily:
        return []
    # Fill in zero-count days so the series is dense (no gaps).
    # This prevents the frontend from truncating charts when some events
    # have sparse data for parts of the date range.
    from datetime import date as _date, timedelta as _td
    d_min = _date.fromisoformat(min(daily))
    d_max = _date.fromisoformat(max(daily))
    result = []
    cur = d_min
    while cur <= d_max:
        result.append({"date": cur.isoformat(), "count": daily.get(cur.isoformat(), 0)})
        cur += _td(days=1)
    return result

async def _topn(project, event, group_by="itemId", n=10):
    data = await _call(f"{CMS_BASE}/{project}/query/top-n",
                       {"event": event, "groupBy": group_by, "n": n})
    if not data: return []
    rows = data.get("top", data) if isinstance(data, dict) else data
    return rows if isinstance(rows, list) else []

async def _timeline(project, user_id, limit=500):
    data = await _call(f"{CMS_BASE}/{project}/query/user-timeline",
                       {"userId": user_id, "since": "30d", "limit": limit})
    if not data: return []
    events = data.get("events", []) if isinstance(data, dict) else data
    return sorted(events, key=lambda e: e.get("timestamp", 0)) if isinstance(events, list) else []

def _etype(e): return e.get("event_type") or e.get("eventType") or ""
def _props(e):
    p = e.get("properties")
    return p if isinstance(p, dict) else {}

def _dense_series_from_day_counts(day_counts: dict) -> list:
    """Convert sparse per-day counts into dense daily series."""
    if not day_counts:
        return []
    from datetime import date as _date, timedelta as _td
    d_min = _date.fromisoformat(min(day_counts))
    d_max = _date.fromisoformat(max(day_counts))
    out = []
    cur = d_min
    while cur <= d_max:
        iso = cur.isoformat()
        out.append({"date": iso, "count": int(day_counts.get(iso, 0) or 0)})
        cur += _td(days=1)
    return out

def _label_from_slug(slug: str) -> str:
    s = (slug or "").strip()
    if not s:
        return "Unknown"
    return s.replace("-", " ").replace("_", " ").title()

def _build_sessions_from_events(events: list) -> dict:
    """Compute compact sessions summary from stored timeline events."""
    if not events:
        return {
            "total_sessions": 0,
            "avg_seconds": 0,
            "median_seconds": 0,
            "p90_seconds": 0,
            "pct_with_logout": 0,
            "depth_distribution": {},
            "activity_breakdown": [],
            "daily_avg": [],
            "hour_distribution": {},
            "note": "No session data available in Supabase yet",
        }

    sessions_map = defaultdict(list)
    for ev in events:
        uid = ev.get("user_id") or ""
        sid = ev.get("session_id") or "nosession"
        ts = int(ev.get("ts") or 0)
        if not ts:
            continue
        sessions_map[f"{uid}::{sid}"].append(ev)

    durations = []
    session_depths = []
    daily_map = defaultdict(list)
    hour_map = defaultdict(int)
    event_type_times = defaultdict(list)
    labels = {
        "article.viewed": "Reading",
        "search.performed": "Searching",
        "video.watched": "Watching",
        "category.viewed": "Browsing",
    }

    for _, evs in sessions_map.items():
        evs_s = sorted(evs, key=lambda e: int(e.get("ts") or 0))
        ts_vals = [int(e.get("ts") or 0) for e in evs_s if int(e.get("ts") or 0)]
        if len(ts_vals) < 2:
            continue

        dur = min((ts_vals[-1] - ts_vals[0]) / 1000, 28800)
        if dur < 5:
            continue
        durations.append(dur)
        session_depths.append(len(evs_s))

        d = datetime.utcfromtimestamp(ts_vals[0] / 1000).strftime("%Y-%m-%d")
        daily_map[d].append(round(dur))
        h = datetime.utcfromtimestamp(ts_vals[0] / 1000).hour
        hour_map[h] += 1

        for i, ev in enumerate(evs_s[:-1]):
            t0 = int(ev.get("ts") or 0)
            t1 = int(evs_s[i+1].get("ts") or 0)
            gap = (t1 - t0) / 1000
            if 2 <= gap <= 600:
                event_type_times[_etype(ev)].append(gap)

    n = len(durations)
    if n == 0:
        return {
            "total_sessions": 0,
            "avg_seconds": 0,
            "median_seconds": 0,
            "p90_seconds": 0,
            "pct_with_logout": 0,
            "depth_distribution": {},
            "activity_breakdown": [],
            "daily_avg": [],
            "hour_distribution": {},
            "note": "No qualifying sessions found in Supabase",
        }

    ds = sorted(durations)
    depth_dist = Counter("bounce" if d <= 2 else "normal" if d <= 9 else "deep" for d in session_depths)
    daily_avg = sorted(
        [{"date": d, "avg_seconds": round(sum(v) / len(v))} for d, v in daily_map.items()],
        key=lambda x: x["date"],
    )
    total_act = sum(sum(v) for k, v in event_type_times.items() if k in labels and v)
    activity_breakdown = []
    for k, lab in labels.items():
        vals = event_type_times.get(k, [])
        if not vals:
            continue
        activity_breakdown.append({
            "label": lab,
            "avg_seconds": round(sum(vals) / len(vals)),
            "pct_time": round(sum(vals) / max(total_act, 1) * 100),
        })
    activity_breakdown.sort(key=lambda x: -x["avg_seconds"])

    return {
        "total_sessions": n,
        "avg_seconds": round(sum(ds) / n),
        "median_seconds": ds[n // 2],
        "p90_seconds": ds[int(n * 0.9)],
        "pct_with_logout": 0,
        "depth_distribution": dict(depth_dist),
        "activity_breakdown": activity_breakdown,
        "daily_avg": daily_avg,
        "hour_distribution": {str(h): c for h, c in sorted(hour_map.items())},
        "note": f"Recovered from Supabase timeline cache ({n} sessions)",
    }

async def _build_fallback_from_supabase() -> tuple:
    """Build best-effort batch and intel payloads from stored Supabase events."""
    empty_batch = {ev["key"]: {"series": []} for ev in EVENTS}
    empty_intel = {
        "articles": {"articles": [], "total_views": 0, "priority": [], "computed_at": _utcnow().isoformat(), "note": "No CMS data"},
        "search": {"top_queries": [], "zero_result": [], "content_gaps": [], "total_searches": 0, "conversion_rate": 0, "frustration_index": 0, "computed_at": _utcnow().isoformat(), "note": "No CMS data"},
        "categories": {"categories": [], "computed_at": _utcnow().isoformat()},
        "videos": {"videos": [], "computed_at": _utcnow().isoformat(), "note": "No CMS data"},
        "sessions": {"total_sessions": 0, "avg_seconds": 0, "median_seconds": 0, "p90_seconds": 0, "pct_with_logout": 0, "depth_distribution": {}, "activity_breakdown": [], "daily_avg": [], "note": "No CMS data"},
        "insights": {"consumption_ratio": 0, "frustration_index": 0, "engagement_velocity": 0, "self_service_rate": 0, "hour_distribution": {}, "weekly_digest": "CMS unavailable — using Supabase fallback", "computed_at": _utcnow().isoformat()},
    }
    if not _sb_enabled:
        return empty_batch, empty_intel, 0, 0

    content_events = await _sb_get_all_events(project="cs-portal-content-events")
    feedback_events = await _sb_get_all_events(project="cs-portal-feedback-events")
    if not content_events and not feedback_events:
        return empty_batch, empty_intel, 0, 0

    # Build batch series for content events available in Supabase.
    day_by_event = defaultdict(lambda: defaultdict(int))
    for ev in content_events:
        et = ev.get("event_type") or ""
        day = ev.get("event_date") or ""
        if not day:
            continue
        if et in ("article.viewed", "search.performed", "video.watched", "category.viewed"):
            day_by_event[et][day] += 1

    batch = {ev["key"]: {"series": []} for ev in EVENTS}
    for key, counts in day_by_event.items():
        batch[key] = {"series": _dense_series_from_day_counts(counts)}

    # Build intel snapshots from Supabase counters.
    article_views = Counter()
    category_views = Counter()
    video_views = Counter()
    queries = Counter()
    for ev in content_events:
        et = ev.get("event_type") or ""
        props = _props(ev)
        item = (ev.get("item_id") or "").strip()
        if et == "article.viewed" and item:
            article_views[item] += 1
        elif et == "category.viewed":
            key = item or (props.get("category") or "")
            if key:
                category_views[key] += 1
        elif et == "video.watched":
            title = (props.get("videoTitle") or item or "").strip()
            if title:
                video_views[title] += 1
        elif et == "search.performed":
            q = (props.get("query") or "").strip().lower()
            if q:
                queries[q] += 1

    helpful = Counter()
    not_helpful = Counter()
    for ev in feedback_events:
        if (ev.get("event_type") or "") != "article.feedback":
            continue
        props = _props(ev)
        aid = (ev.get("item_id") or props.get("articleKey") or "").strip()
        if not aid:
            continue
        v = str(props.get("value") or "").strip().lower()
        if v == "helpful":
            helpful[aid] += 1
        elif v == "not_helpful":
            not_helpful[aid] += 1

    total_views = sum(article_views.values())
    articles = []
    for aid, views in article_views.most_common(20):
        hp = helpful.get(aid, 0)
        nh = not_helpful.get(aid, 0)
        fb_total = hp + nh
        hp_pct = round(hp / fb_total * 100, 1) if fb_total else None
        score_f = 10 if fb_total == 0 else min(30, fb_total * 5)
        if hp_pct is not None and hp_pct >= 80:
            score_f = min(30, score_f + 5)
        if hp_pct is not None and hp_pct < 40:
            score_f = max(0, score_f - 10)
        health = min(100, min(40, round(views / max(total_views, 1) * 400)) + 10 + score_f)
        articles.append({
            "id": aid,
            "label": _label_from_slug(aid),
            "views": int(views),
            "share_pct": round(views / max(total_views, 1) * 100, 1),
            "helpful": int(hp),
            "not_helpful": int(nh),
            "helpful_pct": hp_pct,
            "total_feedback": int(fb_total),
            "avg_seconds": None,
            "min_seconds": None,
            "max_seconds": None,
            "bounce_rate": None,
            "revisits": 0,
            "time_sample": 0,
            "health_score": int(health),
            "needs_attention": bool(views > 50 and (fb_total == 0 or (hp_pct is not None and hp_pct < 50))),
            "is_dead_end": False,
            "next_articles": [],
        })

    sessions = _build_sessions_from_events(content_events)
    search_total = sum(queries.values())
    auth_series = batch.get("auth.login", {}).get("series", [])
    login_total = sum(p.get("count", 0) for p in auth_series)
    article_total = sum(p.get("count", 0) for p in batch.get("article.viewed", {}).get("series", []))
    search_total_series = sum(p.get("count", 0) for p in batch.get("search.performed", {}).get("series", []))
    insights_hour = sessions.get("hour_distribution", {})

    intel = {
        "articles": {
            "articles": articles,
            "total_views": total_views,
            "priority": [{"id": a["id"], "label": a["label"], "views": a["views"], "issue": "No feedback despite views", "health_score": a["health_score"]}
                         for a in articles if a["needs_attention"]][:10],
            "computed_at": _utcnow().isoformat(),
            "note": "Recovered from Supabase timeline cache while CMS is unavailable",
        },
        "search": {
            "top_queries": [{"query": q, "count": c} for q, c in queries.most_common(20)],
            "zero_result": [],
            "content_gaps": [],
            "total_searches": int(search_total),
            "conversion_rate": 0,
            "frustration_index": round(search_total_series / max(login_total, 1) * 100) / 100 if login_total else 0,
            "computed_at": _utcnow().isoformat(),
            "note": "Recovered from Supabase timeline cache",
        },
        "categories": {
            "categories": [{"path": k, "label": _label_from_slug(k.replace("/category/", "")), "count": int(v)}
                           for k, v in category_views.most_common(20)],
            "computed_at": _utcnow().isoformat(),
        },
        "videos": {
            "videos": [{"title": t, "count": int(c), "url": ""} for t, c in video_views.most_common(20)],
            "computed_at": _utcnow().isoformat(),
            "note": "Recovered from Supabase timeline cache",
        },
        "sessions": {
            **sessions,
            "computed_at": _utcnow().isoformat(),
        },
        "insights": {
            "consumption_ratio": round(article_total / max(login_total, 1) * 100) / 100 if login_total else 0,
            "frustration_index": round(search_total_series / max(login_total, 1) * 100) / 100 if login_total else 0,
            "engagement_velocity": 0,
            "self_service_rate": 0,
            "hour_distribution": insights_hour,
            "weekly_digest": "CMS unavailable — showing data rebuilt from Supabase timeline cache",
            "computed_at": _utcnow().isoformat(),
        },
    }

    batch_points = sum(len(v.get("series", [])) for v in batch.values())
    intel_rows = len(articles) + len(video_views) + len(queries) + len(category_views)
    return batch, intel, batch_points, intel_rows

# ── FULL REFRESH — sequential CMS calls with 1s gap each ─────────────────────
async def full_refresh():
    """
    15 fixed CMS calls + up to 12 timeline calls (Supabase) = up to 27 total.
    1s gap between every call. Protected by lock — only one refresh at a time.
    """
    lock = get_lock()
    if lock.locked():
        print("[refresh] already running, skipping", flush=True)
        return
    async with lock:
        print(f"[refresh] starting at {_utcnow().isoformat()}", flush=True)
        old_batch, _ = cache_get("batch:all")
        old_intel, _ = cache_get("intel:all")
        batch  = {}
        intel  = {}
        batch_non_empty_events = 0

        # ── Calls 1-9: time-series for KPIs ──────────────────────────────────
        for ev in EVENTS:
            series = await _timeseries(ev["project"], ev["key"])
            if series:
                batch_non_empty_events += 1
            batch[ev["key"]] = {"series": series}

        if batch_non_empty_events == 0 and old_batch:
            # Protect against upstream outages returning empty responses for all KPIs.
            batch = old_batch
            print("[refresh] CMS returned empty KPI batch — keeping previous cache", flush=True)
        elif batch_non_empty_events == 0 and _sb_enabled:
            # Fresh deploy + CMS outage: recover what we can from already stored timelines.
            sb_batch, _, sb_points, _ = await _build_fallback_from_supabase()
            if sb_points > 0:
                batch = sb_batch
                print(f"[refresh] CMS KPI batch empty — recovered {sb_points} points from Supabase", flush=True)
                cache_set("batch:all", batch)
            else:
                cache_set("batch:all", batch)
        else:
            cache_set("batch:all", batch)
        print(f"[refresh] batch done — {sum(len(v['series']) for v in batch.values())} points", flush=True)

        # ── Call 10: top articles ─────────────────────────────────────────────
        art_rows = await _topn("cs-portal-content-events", "article.viewed", "itemId", 10)

        # ── Call 11: top feedback ─────────────────────────────────────────────
        fb_rows = await _topn("cs-portal-feedback-events", "article.feedback", "itemId", 10)

        # ── Call 12: top categories ───────────────────────────────────────────
        cat_rows = await _topn("cs-portal-content-events", "category.viewed", "itemId", 10)

        # ── Calls 13-15: get all available user IDs from 3 sources ─────────────
        video_user_rows  = await _topn("cs-portal-content-events", "video.watched",    "userId", 10)
        search_user_rows = await _topn("cs-portal-content-events", "search.performed", "userId", 10)
        art_user_rows    = await _topn("cs-portal-content-events", "article.viewed",   "userId", 10)

        # Collect all unique non-anonymous user IDs
        seen = set()
        all_known_users = []
        for rows in (video_user_rows, search_user_rows, art_user_rows):
            for r in rows:
                uid = r.get("userId") or r.get("itemId") or ""
                if uid and uid.lower() not in ("anonymous", "") and uid not in seen:
                    seen.add(uid)
                    all_known_users.append(uid)

        # With Supabase: rotate through all known users across refresh cycles
        # Without Supabase: fall back to top 4
        if _sb_enabled:
            users_to_fetch = await _sb_get_unfetched_users(all_known_users, limit=10)
        else:
            users_to_fetch = all_known_users[:4]

        print(f"[refresh] fetching timelines for {len(users_to_fetch)} users (of {len(all_known_users)} known)", flush=True)

        # ── Content timelines ────────────────────────────────────────────────
        all_events = []
        fb_events  = []
        for uid in users_to_fetch:
            tl = await _timeline("cs-portal-content-events", uid, limit=500)
            all_events.extend(tl)
            print(f"[refresh] timeline {uid[:16]}: {len(tl)} events", flush=True)
            if _sb_enabled:
                await _sb_store_events(uid, "cs-portal-content-events", tl)

        # Feedback timelines for first 2 users
        for uid in users_to_fetch[:2]:
            tl = await _timeline("cs-portal-feedback-events", uid, limit=200)
            fb_events.extend(tl)
            if _sb_enabled:
                await _sb_store_events(uid, "cs-portal-feedback-events", tl)
            print(f"[refresh] fb-timeline {uid[:16]}: {len(tl)} events", flush=True)

        print(f"[refresh] intel calls done — {len(all_events)} content events, {len(fb_events)} feedback events", flush=True)

        sorted_ev = sorted(all_events, key=lambda e: e.get("timestamp", 0))

        # ── Extract feedback sentiment from feedback timelines ────────────────
        fb_helpful     = defaultdict(int)   # aid -> helpful count
        fb_not_helpful = defaultdict(int)   # aid -> not_helpful count
        for ev in fb_events:
            if _etype(ev) != "article.feedback":
                continue
            props = _props(ev)
            aid   = ev.get("item_id") or props.get("articleKey") or ""
            val   = props.get("value", "")
            if not aid:
                continue
            if val == "helpful":
                fb_helpful[aid] += 1
            elif val == "not_helpful":
                fb_not_helpful[aid] += 1

        # ── Extract videos ────────────────────────────────────────────────────
        video_counter = Counter()
        video_urls    = {}
        for ev in all_events:
            if _etype(ev) == "video.watched":
                props = _props(ev)
                title = props.get("videoTitle", "").strip()
                url   = props.get("videoUrl", "")
                if title:
                    video_counter[title] += 1
                    if url and title not in video_urls:
                        video_urls[title] = url

        # ── Extract search queries ────────────────────────────────────────────
        query_counter = Counter()
        zero_result_q = Counter()
        search_total  = 0
        search_conv   = 0
        for i, ev in enumerate(sorted_ev):
            if _etype(ev) != "search.performed": continue
            props = _props(ev)
            q = (props.get("query") or "").strip()
            if not q: continue
            ts = ev.get("timestamp", 0)
            is_final = True
            if i+1 < len(sorted_ev):
                nev = sorted_ev[i+1]
                if _etype(nev) == "search.performed" and (nev.get("timestamp",0)-ts) < SEARCH_GAP_MS:
                    is_final = False
            if not is_final: continue
            search_total += 1
            ql = q.lower()
            query_counter[ql] += 1
            if props.get("resultCount") == 0:
                zero_result_q[ql] += 1
            sid = ev.get("session_id") or ""
            for j in range(i+1, len(sorted_ev)):
                nev = sorted_ev[j]
                if nev.get("session_id") != sid and sid: break
                if _etype(nev) == "article.viewed": search_conv += 1; break

        # ── Extract article intelligence from timelines ───────────────────────
        article_times   = defaultdict(list)   # aid -> [seconds spent]
        article_next    = defaultdict(Counter) # aid -> {next_aid: count}
        article_bounces = defaultdict(int)     # aid -> bounce count (<30s)
        article_revisits= defaultdict(int)     # aid -> revisit count
        session_durations = []
        session_depths    = []
        hour_counts       = defaultdict(int)   # hour -> event count
        seen_article_sessions = defaultdict(set)  # aid -> set of session_ids

        # Group by session
        sessions_map = defaultdict(list)
        for ev in sorted_ev:
            sid = ev.get("session_id") or "nosession"
            sessions_map[sid].append(ev)

        for sid, evts in sessions_map.items():
            if sid == "nosession": continue
            evts_s = sorted(evts, key=lambda e: e.get("timestamp", 0))
            ts_list = [e.get("timestamp",0) for e in evts_s if e.get("timestamp")]
            if len(ts_list) >= 2:
                dur = min((max(ts_list)-min(ts_list))/1000, 28800)
                session_durations.append(round(dur))
                # Hour of day from session start
                hour = datetime.utcfromtimestamp(min(ts_list)/1000).hour
                hour_counts[hour] += 1

            meaningful = [e for e in evts_s if _etype(e) != "profile.viewed"]
            session_depths.append(len(meaningful))

            # Article time spent + bounce + navigation paths
            for i, ev in enumerate(evts_s):
                et    = _etype(ev)
                props = _props(ev)
                aid   = ev.get("item_id") or props.get("articleKey") or ""
                if et == "article.viewed" and aid:
                    t0 = ev.get("timestamp", 0)
                    # Revisit detection
                    if sid in seen_article_sessions[aid]:
                        article_revisits[aid] += 1
                    else:
                        seen_article_sessions[aid].add(sid)
                    # Time spent + bounce
                    for j in range(i+1, len(evts_s)):
                        nev = evts_s[j]
                        if _etype(nev) == "article.viewed":
                            gap = (nev.get("timestamp",0) - t0) / 1000
                            if 5 <= gap <= 300:
                                article_times[aid].append(round(gap))
                                if gap < 30:
                                    article_bounces[aid] += 1
                            next_aid = nev.get("item_id") or _props(nev).get("articleKey") or ""
                            if next_aid and next_aid != aid:
                                article_next[aid][next_aid] += 1
                            break

        # ── Session analytics from timelines ──────────────────────────────────
        n_sess = len(session_durations)
        if n_sess > 0:
            durs_s = sorted(session_durations)
            avg_dur    = round(sum(durs_s)/n_sess)
            median_dur = durs_s[n_sess//2]
            p90_dur    = durs_s[int(n_sess*0.9)]
            depths     = Counter("bounce" if d<=2 else "normal" if d<=9 else "deep" for d in session_depths)
        else:
            avg_dur = median_dur = p90_dur = 0
            depths = {}

        # ── daily_avg: avg session duration grouped by date ───────────────────
        daily_dur_map = defaultdict(list)  # date -> [seconds]
        for sid, evts in sessions_map.items():
            if sid == "nosession": continue
            evts_s = sorted(evts, key=lambda e: e.get("timestamp", 0))
            ts_list = [e.get("timestamp",0) for e in evts_s if e.get("timestamp")]
            if len(ts_list) >= 2:
                dur = min((max(ts_list)-min(ts_list))/1000, 28800)
                d   = datetime.utcfromtimestamp(min(ts_list)/1000).strftime("%Y-%m-%d")
                daily_dur_map[d].append(round(dur))
        daily_avg = [
            {"date": d, "avg_seconds": round(sum(v)/len(v))}
            for d, v in sorted(daily_dur_map.items())
        ]

        # ── activity_breakdown: avg seconds per event type per session ────────
        event_type_times = defaultdict(list)
        for sid, evts in sessions_map.items():
            if sid == "nosession": continue
            evts_s = sorted(evts, key=lambda e: e.get("timestamp", 0))
            # Time between consecutive events of each type
            for i, ev in enumerate(evts_s[:-1]):
                t0 = ev.get("timestamp", 0)
                t1 = evts_s[i+1].get("timestamp", 0)
                gap = (t1 - t0) / 1000
                if 2 <= gap <= 600:  # ignore sub-2s and >10min gaps
                    event_type_times[_etype(ev)].append(gap)
        activity_keys = ["article.viewed", "search.performed", "video.watched", "category.viewed"]
        activity_labels = {"article.viewed":"Reading","search.performed":"Searching",
                           "video.watched":"Watching","category.viewed":"Browsing"}
        total_act_time = sum(
            sum(event_type_times[k]) for k in activity_keys if event_type_times[k]
        )
        activity_breakdown = []
        for k in activity_keys:
            times_k = event_type_times.get(k, [])
            if not times_k: continue
            avg_s = round(sum(times_k)/len(times_k))
            pct   = round(sum(times_k)/max(total_act_time,1)*100)
            activity_breakdown.append({"label": activity_labels[k], "avg_seconds": avg_s, "pct_time": pct})
        activity_breakdown.sort(key=lambda x: -x["avg_seconds"])

        # ── pct_with_logout: sessions that have an explicit logout ────────────
        sessions_with_logout = sum(
            1 for sid, evts in sessions_map.items()
            if sid != "nosession" and any(_etype(e) == "auth.logout" for e in evts)
        )
        pct_logout = round(sessions_with_logout / max(n_sess, 1) * 100, 1)

        # ── KPI-derived metrics (from batch — zero extra calls) ───────────────
        def series_total(key):
            return sum(p.get("count",0) for p in batch.get(key,{}).get("series",[]))
        def series_last_n(key, n=7):
            s = batch.get(key,{}).get("series",[])
            return sum(p.get("count",0) for p in s[-n:]) if s else 0

        total_logins   = series_total("auth.login")
        total_articles = series_total("article.viewed")
        total_searches = series_total("search.performed")
        total_videos   = series_total("video.watched")

        # Content consumption ratio — articles per login
        consumption_ratio = round(total_articles/max(total_logins,1)*100)/100

        # Search frustration index — searches per login (high = struggling)
        frustration_idx = round(total_searches/max(total_logins,1)*100)/100

        # Engagement velocity — last 7 days vs previous 7 days
        logins_last7  = series_last_n("auth.login", 7)
        logins_prev7  = series_last_n("auth.login", 14) - logins_last7
        velocity = round((logins_last7 - logins_prev7) / max(logins_prev7,1) * 100, 1) if logins_prev7 else 0

        # Self-service sessions (login → article → no search) from timelines
        self_service = 0
        total_sess_with_login = 0
        for sid, evts in sessions_map.items():
            types = [_etype(e) for e in evts]
            if "auth.login" in types:
                total_sess_with_login += 1
                if "article.viewed" in types and "search.performed" not in types:
                    self_service += 1
        self_service_rate = round(self_service/max(total_sess_with_login,1)*100,1)

        # ── Assemble articles with full intelligence ──────────────────────────
        fb_map      = {r.get("itemId") or "": int(r.get("count",0)) for r in fb_rows}
        total_views = sum(int(r.get("count",0)) for r in art_rows)
        articles    = []
        for r in art_rows:
            aid   = r.get("itemId") or r.get("item_id") or ""
            views = int(r.get("count", 0))
            fb    = fb_map.get(aid, 0)
            times = article_times.get(aid, [])
            avg_t = round(sum(times)/len(times)) if times else None
            min_t = round(min(times)) if times else None
            max_t = round(max(times)) if times else None
            bounce_rate = round(article_bounces.get(aid,0)/max(len(times),1)*100,1) if times else None
            revisits = article_revisits.get(aid, 0)
            next_arts = [{"id":k,"label":k.replace("-"," ").title(),"count":v}
                         for k,v in sorted(article_next.get(aid,{}).items(),key=lambda x:-x[1])[:3]]
            # Feedback sentiment from timeline sample
            hlp     = fb_helpful.get(aid, 0)
            not_hlp = fb_not_helpful.get(aid, 0)
            hlp_total = hlp + not_hlp
            hlp_pct = round(hlp / hlp_total * 100, 1) if hlp_total > 0 else None
            # Use timeline sentiment if available, else fall back to top-n count only
            fb_total = hlp_total if hlp_total > 0 else fb
            # Health score: views(40) + time_spent(30) + feedback(30)
            score_v = min(40, round(views/max(total_views,1)*400))
            score_t = min(30, round(avg_t/300*30)) if avg_t else 10
            score_f = min(30, fb_total * 5) if fb_total > 0 else 10
            # Bonus for high helpful rate
            if hlp_pct and hlp_pct >= 80: score_f = min(30, score_f + 5)
            # Penalty for high bounce or low helpful rate
            if bounce_rate and bounce_rate > 60: score_t = max(0, score_t - 15)
            if hlp_pct and hlp_pct < 40: score_f = max(0, score_f - 10)
            health = score_v + score_t + score_f
            # Priority flag: high views + high bounce or no feedback or low helpful rating
            needs_attention = (views > 50 and (
                (bounce_rate and bounce_rate > 60)
                or fb_total == 0
                or (hlp_pct is not None and hlp_pct < 50)
            ))
            articles.append({
                "id": aid,
                "label": aid.replace("-"," ").replace("_"," ").title(),
                "views": views,
                "share_pct": round(views/total_views*100,1) if total_views else 0,
                "helpful": hlp, "not_helpful": not_hlp, "helpful_pct": hlp_pct,
                "total_feedback": fb_total,
                "avg_seconds": avg_t,
                "min_seconds": min_t,
                "max_seconds": max_t,
                "bounce_rate": bounce_rate,
                "revisits": revisits,
                "time_sample": len(times),
                "health_score": health,
                "needs_attention": needs_attention,
                "is_dead_end": len(article_next) > 0 and aid not in {k for c in article_next.values() for k in c},
                "next_articles": next_arts,
            })

        # Sort by needs_attention first, then by views
        articles.sort(key=lambda x: (-x["needs_attention"], -x["views"]))

        # ── Article improvement priority ──────────────────────────────────────
        priority = [
            {"id": a["id"], "label": a["label"], "views": a["views"],
             "issue": "High bounce rate" if (a["bounce_rate"] and a["bounce_rate"]>60)
                      else "No feedback despite views" if a["total_feedback"]==0 and a["views"]>50
                      else "Low time spent" if (a["avg_seconds"] and a["avg_seconds"]<30)
                      else "Watch",
             "health_score": a["health_score"]}
            for a in articles if a["needs_attention"]
        ]

        # ── Content gaps ──────────────────────────────────────────────────────
        art_slugs = {(r.get("itemId") or "").lower().replace("-"," ") for r in art_rows}
        gaps = []
        for q, cnt in query_counter.most_common(20):
            q_words = set(q.lower().split())
            matched = any(len(q_words & set(s.split())) >= 1 for s in art_slugs)
            gaps.append({"query":q,"count":cnt,
                         "is_zero_result":q in zero_result_q,
                         "has_content":matched})

        # ── Weekly digest ─────────────────────────────────────────────────────
        top_art = articles[0]["label"] if articles else "N/A"
        digest_items = []
        if velocity > 10:  digest_items.append(f"Logins up {velocity}% vs last week")
        elif velocity < -10: digest_items.append(f"Logins down {abs(velocity)}% vs last week — worth investigating")
        if priority:       digest_items.append(f"{len(priority)} article(s) need attention")
        if gaps and any(not g["has_content"] for g in gaps):
            n_gaps = sum(1 for g in gaps if not g["has_content"])
            digest_items.append(f"{n_gaps} search term(s) have no matching article")
        digest_items.append(f"Most viewed: {top_art}")
        digest = " · ".join(digest_items) if digest_items else "No significant changes this week"

        # ── Store intel ───────────────────────────────────────────────────────
        intel = {
            "articles": {
                "articles": articles, "total_views": total_views,
                "priority": priority,
                "computed_at": _utcnow().isoformat(),
                "note": "Views & feedback from top-n. Time/bounce from timeline sample.",
            },
            "search": {
                "top_queries":     [{"query":q,"count":c} for q,c in query_counter.most_common(20)],
                "zero_result":     [{"query":q,"count":c} for q,c in zero_result_q.most_common(10)],
                "content_gaps":    gaps,
                "total_searches":  search_total,
                "conversion_rate": round(search_conv/search_total*100,1) if search_total else 0,
                "frustration_index": frustration_idx,
                "computed_at":     _utcnow().isoformat(),
                "note": "Queries from timeline sample. Frustration index = searches/logins.",
            },
            "categories": {
                "categories": [{
                    "path":  r.get("itemId") or "",
                    "label": (r.get("itemId") or "").replace("/category/","").replace("-"," ").title() or "Home",
                    "count": int(r.get("count",0)),
                } for r in cat_rows],
                "computed_at": _utcnow().isoformat(),
            },
            "videos": {
                "videos": [{"title":t,"count":c,"url":video_urls.get(t,"")}
                           for t,c in video_counter.most_common(20)],
                "computed_at": _utcnow().isoformat(),
                "note": "From timeline sample.",
            },
            "sessions": {
                "total_sessions":  n_sess,
                "avg_seconds":     avg_dur,
                "median_seconds":  median_dur,
                "p90_seconds":     p90_dur,
                "pct_with_logout": pct_logout,
                "depth_distribution": dict(depths),
                "activity_breakdown": activity_breakdown,
                "daily_avg": daily_avg,
                "computed_at": _utcnow().isoformat(),
                "note": "From timeline sample of top video/search users.",
            },
            "insights": {
                "consumption_ratio":  consumption_ratio,
                "frustration_index":  frustration_idx,
                "engagement_velocity":velocity,
                "self_service_rate":  self_service_rate,
                "hour_distribution":  dict(hour_counts),
                "weekly_digest":      digest,
                "computed_at":        _utcnow().isoformat(),
            },
        }
        has_intel_payload = any([
            len(articles) > 0,
            len(video_counter) > 0,
            search_total > 0,
            n_sess > 0,
            len(cat_rows) > 0,
        ])
        if not has_intel_payload and old_intel:
            # Keep the last known-good snapshot if CMS/timeline calls failed upstream.
            intel = old_intel
            print("[refresh] CMS returned empty intel payload — keeping previous cache", flush=True)
        elif not has_intel_payload and _sb_enabled:
            # Fresh deploy + CMS outage: build a best-effort intel payload from Supabase.
            _, sb_intel, _, sb_rows = await _build_fallback_from_supabase()
            if sb_rows > 0:
                intel = sb_intel
                print(f"[refresh] CMS intel payload empty — recovered fallback intel from Supabase ({sb_rows} rows)", flush=True)
                cache_set("intel:all", intel)
            else:
                cache_set("intel:all", intel)
        else:
            cache_set("intel:all", intel)
        print(f"[refresh] complete — {len(articles)} articles, {len(video_counter)} videos, {search_total} searches", flush=True)

# ── Background loop — one refresh every 2 hours ────────────────────────────────
# Supabase free tier pauses after 1 week of inactivity.
# We ping it every 6 hours to keep it alive — cheap and reliable.
SUPABASE_KEEPALIVE_SEC = 6 * 3600

async def _supabase_keepalive_loop():
    """Ping Supabase every 6 hours to prevent free-tier pausing."""
    await asyncio.sleep(60)  # wait for startup to settle
    while True:
        if _sb_enabled:
            try:
                result = await _sb_request("GET", "/cs_user_fetch_log?select=user_id&limit=1")
                print(f"[supabase] keepalive ping OK", flush=True)
            except Exception as ex:
                print(f"[supabase] keepalive ping failed: {ex}", flush=True)
        await asyncio.sleep(SUPABASE_KEEPALIVE_SEC)

async def _refresh_loop():
    await asyncio.sleep(10)  # let server boot fully
    while True:
        try:
            await full_refresh()
        except Exception as ex:
            print(f"[loop] error: {ex}", flush=True)
        # Refresh CSAT from Domo on the same 2-hour cycle (no-op if not configured).
        if _domo_configured():
            try:
                await asyncio.to_thread(_refresh_csat_from_domo)
            except Exception as ex:
                print(f"[loop] domo csat refresh error: {ex}", flush=True)
        await asyncio.sleep(REFRESH_SEC)

# ── Lifespan ───────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app):
    load_cache_from_disk()
    _load_csat_csv()
    if _sb_enabled:
        await _sb_ensure_tables()
        user_count = await _sb_get_user_count()
        print(f"[supabase] connected — {user_count} users in DB", flush=True)
    else:
        print("[supabase] not configured — using local timeline only", flush=True)
    t = asyncio.create_task(_refresh_loop())
    k = asyncio.create_task(_supabase_keepalive_loop())  # prevents Supabase free-tier pause
    yield
    t.cancel()
    k.cancel()
    global _client
    if _client and not _client.is_closed:
        await _client.aclose()

# ── App ────────────────────────────────────────────────────────────────────────
app = FastAPI(lifespan=lifespan, title="CS Portal Analytics")
app.add_middleware(GZipMiddleware, minimum_size=500)
_allowed_origins = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "").split(",") if o.strip()]
if _allowed_origins:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=_allowed_origins,
        allow_methods=["GET", "POST", "HEAD", "OPTIONS"],
        allow_headers=["Authorization", "Content-Type", "X-Admin-Token"],
    )

# ── Endpoints ──────────────────────────────────────────────────────────────────
@app.get("/api/config")
async def api_config():
    return {
        "dashboard_version": DASHBOARD_VERSION,
        "environment": APP_ENV,
        "data_start": DATA_START,
        "events": EVENTS,
        "csat": {
            "schema_version": CSAT_SCHEMA_VERSION,
            "key_legend": CSAT_INDEX_KEY_LEGEND,
            "metric_definitions": CSAT_METRIC_DEFINITIONS,
            "admin_token_required": ADMIN_TOKEN_REQUIRED,
            "admin_token_configured": bool(ADMIN_TOKEN),
            "raw_drilldowns_public": CSAT_RAW_PUBLIC,
            "refresh_requires_admin": CSAT_REFRESH_REQUIRES_ADMIN,
            "team_filter_enabled": bool(CSAT_TEAMS_ALLOW),
        },
    }

@app.get("/api/metrics/batch")
async def batch_metrics():
    data, _ = cache_get("batch:all")
    if not data:
        data = {}
    payload  = json.dumps(data, separators=(",",":"))
    etag_val = hashlib.md5(payload.encode()).hexdigest()[:16]
    return Response(content=payload, media_type="application/json",
        headers={"Cache-Control":f"public, max-age={CACHE_TTL}, stale-while-revalidate={STALE_TTL}",
                 "ETag":f'"{etag_val}"',"Vary":"Accept-Encoding"})

@app.get("/api/articles")
async def api_articles():
    data, _ = cache_get("intel:all")
    return (data or {}).get("articles", {"articles":[],"total_views":0})

@app.get("/api/search")
async def api_search():
    data, _ = cache_get("intel:all")
    return (data or {}).get("search", {"top_queries":[],"zero_result":[],"content_gaps":[],"total_searches":0,"conversion_rate":0})

@app.get("/api/sessions")
async def api_sessions():
    data, _ = cache_get("intel:all")
    return (data or {}).get("sessions", {"total_sessions":0,"note":"Session sample from timelines."})

@app.get("/api/insights")
async def api_insights():
    data, _ = cache_get("intel:all")
    return (data or {}).get("insights", {
        "consumption_ratio":0,"frustration_index":0,
        "engagement_velocity":0,"self_service_rate":0,
        "hour_distribution":{},"weekly_digest":"No data yet.","computed_at":""
    })

@app.get("/api/videos")
async def api_videos():
    data, _ = cache_get("intel:all")
    return (data or {}).get("videos", {"videos":[]})

@app.get("/api/categories")
async def api_categories():
    data, _ = cache_get("intel:all")
    return (data or {}).get("categories", {"categories":[]})

@app.get("/api/sessions/full")
async def api_sessions_full(date_from: str = "", date_to: str = ""):
    """Full session analytics from all users stored in Supabase.
    Falls back to in-memory timeline sample if Supabase is unavailable.
    """
    if not _sb_enabled:
        # Supabase not configured — serve in-memory sample so dashboard still works
        data, _ = cache_get("intel:all")
        sessions = (data or {}).get("sessions", {})
        sessions["note"] = "Supabase not configured — showing in-memory sample (top users only)"
        return {"available": True, **sessions}

    try:
        events = await _sb_get_all_events(date_from=date_from, date_to=date_to)
    except Exception as ex:
        # Supabase is down — fall back to in-memory cache so dashboard stays functional
        print(f"[supabase] sessions/full fallback to cache: {ex}", flush=True)
        data, _ = cache_get("intel:all")
        sessions = (data or {}).get("sessions", {})
        sessions["note"] = "Supabase temporarily unavailable — showing cached sample"
        return {"available": True, **sessions}

    if not events:
        user_count = await _sb_get_user_count()
        return {"available": True, "total_users": user_count, "total_sessions": 0,
                "note": f"{user_count} users in DB — building up data each refresh cycle"}

    from datetime import datetime as _dt
    sessions_map = defaultdict(list)
    for ev in events:
        uid = ev.get("user_id", "")
        sid = ev.get("session_id") or ""
        ts  = ev.get("ts", 0)
        if not ts: continue
        sessions_map[f"{uid}::{sid or 'nosession'}"].append(ts)

    durations, depths = [], []
    daily_map, hour_map = defaultdict(list), defaultdict(int)
    for key, timestamps in sessions_map.items():
        if len(timestamps) < 2: continue
        ts_s = sorted(timestamps)
        dur  = min((ts_s[-1] - ts_s[0]) / 1000, 28800)
        if dur < 5: continue
        durations.append(dur)
        depths.append(len(timestamps))
        d = _dt.utcfromtimestamp(ts_s[0] / 1000).strftime("%Y-%m-%d")
        daily_map[d].append(round(dur))
        hour_map[_dt.utcfromtimestamp(ts_s[0] / 1000).hour] += 1

    n = len(durations)
    if n == 0:
        return {"available": True, "total_sessions": 0, "note": "No qualifying sessions found"}

    ds = sorted(durations)
    depth_dist = Counter("bounce" if d<=2 else "normal" if d<=9 else "deep" for d in depths)
    daily_avg  = sorted([{"date":d,"avg_seconds":round(sum(v)/len(v))} for d,v in daily_map.items()], key=lambda x:x["date"])

    # Activity breakdown
    event_by_sess = defaultdict(list)
    for ev in events:
        uid = ev.get("user_id",""); sid = ev.get("session_id") or ""
        event_by_sess[f"{uid}::{sid or 'nosession'}"].append(ev)
    event_times = defaultdict(list)
    for key, evs in event_by_sess.items():
        evs_s = sorted(evs, key=lambda e: e.get("ts",0))
        for i, ev in enumerate(evs_s[:-1]):
            gap = (evs_s[i+1].get("ts",0) - ev.get("ts",0)) / 1000
            if 2 <= gap <= 600:
                event_times[ev.get("event_type","")].append(gap)
    labels = {"article.viewed":"Reading","search.performed":"Searching","video.watched":"Watching","category.viewed":"Browsing"}
    total_t = sum(sum(v) for v in event_times.values() if v)
    activity_breakdown = sorted([
        {"label":labels[k],"avg_seconds":round(sum(v)/len(v)),"pct_time":round(sum(v)/max(total_t,1)*100)}
        for k,v in event_times.items() if k in labels and v
    ], key=lambda x:-x["avg_seconds"])

    user_count = await _sb_get_user_count()
    return {
        "available": True, "total_users": user_count, "total_sessions": n,
        "avg_seconds": round(sum(ds)/n), "median_seconds": ds[n//2], "p90_seconds": ds[int(n*.9)],
        "pct_with_logout": 0,
        "depth_distribution": dict(depth_dist),
        "daily_avg": daily_avg, "activity_breakdown": activity_breakdown,
        "hour_distribution": {str(h):c for h,c in sorted(hour_map.items())},
        "computed_at": _utcnow().isoformat(),
        "note": f"From {user_count} users in Supabase — {'full' if user_count >= 50 else 'building up coverage'} ({n} sessions)",
    }

def _current_csat_index_data() -> dict:
    """Return the active CSAT index from memory or disk without leaking secrets."""
    index_data = _csat_index.get("index_data")
    if index_data:
        return index_data
    if os.path.exists(CSAT_JSON_PATH):
        try:
            with open(CSAT_JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as exc:
            print(f"[csat] failed to read {CSAT_JSON_PATH}: {exc}", flush=True)
    return {}


def _sanitize_csat_index_for_dashboard(index_data: dict) -> dict:
    """Return the chart-ready CSAT index without call-level summaries/IDs.

    This keeps the dashboard useful for normal users while keeping sensitive
    call summaries, call IDs, response IDs, opportunity IDs, owner IDs, and
    created-by IDs behind /api/csat/raw.
    """
    if not index_data:
        return {"available": False, "note": "CSAT Domo data is not available yet"}
    safe = json.loads(json.dumps(index_data))
    for dm in (safe.get("days") or {}).values():
        for rv in (dm.get("rs") or {}).values():
            rv.pop("sm", None)
        for tv in (dm.get("tm") or {}).values():
            for rv in (tv.get("rs") or {}).values():
                rv.pop("sm", None)
        for cv in (dm.get("cn") or {}).values():
            cv.pop("c", None)
            for rv in (cv.get("rs") or {}).values():
                rv.pop("sm", None)
    safe["access"] = {
        "level": "aggregate",
        "sensitive_fields_removed": True,
        "protected_fields": [
            "call_id", "call_summary", "opportunity_id", "response_id",
            "created_by_id", "owner_id", "consultant_call_cache",
        ],
    }
    return safe


def _csat_no_cache_headers() -> dict:
    return {
        "Cache-Control": "no-store, no-cache, must-revalidate",
        "Pragma": "no-cache",
    }


@app.get("/api/csat/view")
async def api_csat_view():
    """Public aggregate CSAT index used by the dashboard by default.

    It is sourced from the same Domo-built index as /api/csat/raw, but strips
    call-level payloads. The CSAT page now uses /api/csat/raw directly so users
    can drill without an unlock modal.
    """
    index_data = _current_csat_index_data()
    return JSONResponse(_sanitize_csat_index_for_dashboard(index_data), headers=_csat_no_cache_headers())


@app.get("/api/csat/raw")
async def api_csat_raw():
    """Serve the full Domo-backed CSAT index for open in-dashboard drilldowns.

    The product requirement is no locked CSAT drilldown state. This endpoint may
    include call IDs, call summaries, opportunity IDs, and consultant-level call
    caches, so deploy this dashboard only inside the intended internal network.
    """
    index_data = _current_csat_index_data()
    if index_data:
        return JSONResponse(index_data, headers=_csat_no_cache_headers())
    return JSONResponse({"available": False, "note": "CSAT Domo data is not available yet"},
                        headers=_csat_no_cache_headers())


@app.post("/upload/csat")
async def upload_csat(file: UploadFile = File(...), admin_ok: bool = Depends(require_admin_token)):
    """Protected break-glass CSAT ingestion. Normal dashboard flow pulls from Domo.

    This endpoint is intentionally not exposed in the browser UI. It exists for
    emergency recovery/backfill only and requires DASHBOARD_ADMIN_TOKEN.
    """
    # Validate file type
    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="File must be .xlsx, .xls, or .csv")

    try:
        data = await file.read()
        print(f"[csat-breakglass] received {fname} ({len(data)//1024} KB)", flush=True)

        if fname.lower().endswith(".csv"):
            # Parse CSV directly
            rows = _parse_csv_text(data.decode("utf-8", errors="ignore"))
        else:
            rows = _parse_excel_bytes(data)

        if len(rows) < 100:
            raise HTTPException(status_code=400,
                detail=f"Only {len(rows)} valid rows found — check file format")

        # Rebuild index in memory. This endpoint is intentionally retained as
        # a protected break-glass path; the normal CSAT dashboard flow pulls
        # directly from Domo.
        index_data = _build_csat_index(rows)
        _mark_csat_source("manual_break_glass_upload")

        # Save to disk so it persists and gets served via /api/csat/raw
        _save_csat_json()

        return {
            "success":    True,
            "rows":       len(rows),
            "date_min":   index_data["date_min"],
            "date_max":   index_data["date_max"],
            "generated":  index_data["generated"],
            "message":    f"✓ CSAT data updated: {len(rows):,} surveys loaded",
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")


@app.get("/api/csat")
async def api_csat(admin_ok: bool = Depends(require_admin_token)):
    """Return the current pre-indexed CSAT payload using the canonical schema."""
    index_data = _csat_index.get("index_data")
    if index_data:
        return index_data
    if not _csat_index.get("day_map"):
        return {"available": False, "note": "CSAT Domo data is not available yet"}
    return {
        "available": True,
        "schema_version": CSAT_SCHEMA_VERSION,
        "key_legend": CSAT_INDEX_KEY_LEGEND,
        "metric_definitions": CSAT_METRIC_DEFINITIONS,
        "date_min": _csat_index.get("date_min", ""),
        "date_max": _csat_index.get("date_max", ""),
        "days": _csat_index.get("day_map", {}),
        "total_rows": len(_csat_rows),
    }

@app.get("/api/csat/status")
async def api_csat_status():
    """Public, non-sensitive CSAT status and data-quality summary."""
    idx = _csat_index.get("index_data") or {}
    quality = idx.get("quality") or {}
    return {
        "available": bool(idx.get("available")) or bool(_csat_rows),
        "schema_version": idx.get("schema_version", CSAT_SCHEMA_VERSION),
        "date_min": idx.get("date_min", _csat_index.get("date_min", "")),
        "date_max": idx.get("date_max", _csat_index.get("date_max", "")),
        "total_rows": idx.get("total_rows", len(_csat_rows)),
        "raw_total_rows": idx.get("raw_total_rows", len(_csat_rows)),
        "generated": idx.get("generated", ""),
        "source": {
            "active": CSAT_REFRESH_SOURCE.get("source") or ("domo" if _domo_configured() else "bundled_csv"),
            "domo_configured": _domo_configured(),
            "dataset_id_set": bool(DOMO_DATASET_ID),
            "last_success_at": CSAT_REFRESH_SOURCE.get("last_success_at", ""),
            "loaded_at": CSAT_REFRESH_SOURCE.get("loaded_at", ""),
            "last_error": CSAT_REFRESH_SOURCE.get("last_error", ""),
            "refresh_cadence_seconds": REFRESH_SEC,
            "raw_drilldowns_public": CSAT_RAW_PUBLIC,
            "refresh_requires_admin": CSAT_REFRESH_REQUIRES_ADMIN,
        },
        "quality": {
            "indexed_rows": quality.get("indexed_rows"),
            "raw_rows": quality.get("raw_rows"),
            "dropped_by_team_filter": quality.get("dropped_by_team_filter"),
            "distinct_teams": quality.get("distinct_teams"),
            "distinct_consultants": quality.get("distinct_consultants"),
            "distinct_days": quality.get("distinct_days"),
            "true_fcr_available": quality.get("true_fcr_available"),
            "team_filter_enabled": bool(CSAT_TEAMS_ALLOW),
        },
    }

@app.get("/debug/csat")
async def debug_csat(admin_ok: bool = Depends(require_admin_token)):
    """Check whether call_quality.csv was loaded successfully."""
    summary_nonempty = [
        r for r in _csat_rows
        if str(r.get("summary") or "").strip() or str(r.get("summary_raw") or "").strip()
    ]
    paths = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "call_quality.csv"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "call_quality.csv"),
        os.path.join(os.getcwd(), "call_quality.csv"),
        os.path.join(os.getcwd(), "data", "call_quality.csv"),
        "/app/call_quality.csv",
        "/app/data/call_quality.csv",
    ]
    return {
        "rows_loaded": len(_csat_rows),
        "available": len(_csat_rows) > 0,
        "cwd": os.getcwd(),
        "paths_checked": {p: os.path.exists(p) for p in paths},
        "sample": _csat_rows[:2] if _csat_rows else [],
        "summary_rows_nonempty": len(summary_nonempty),
        "summary_sample": summary_nonempty[:2] if summary_nonempty else [],
        "columns_detected": _last_csv_columns,
        "raw_sample_row": _last_raw_sample,
        "domo": {
            "configured": _domo_configured(),
            "client_id_set": bool(DOMO_CLIENT_ID),
            "secret_set": bool(DOMO_CLIENT_SECRET),
            "dataset_id_set": bool(DOMO_DATASET_ID),
            "source": "domo" if _domo_configured() else "bundled_csv",
        },
        "team_allow_list": CSAT_TEAMS_ALLOW,
    }

@app.get("/debug/csat/dates")
async def debug_csat_dates(admin_ok: bool = Depends(require_admin_token)):
    """Per-month row counts — reveals gaps (e.g. a missing March–April) and whether
    a gap is in the raw data vs introduced downstream."""
    from collections import Counter
    by_month = Counter()
    by_day = Counter()
    for r in _csat_rows:
        d = (r.get("date") or "")[:10]
        if len(d) >= 7:
            by_month[d[:7]] += 1
        if d:
            by_day[d] += 1
    months = dict(sorted(by_month.items()))
    days_sorted = sorted(by_day.keys())
    # find gaps of >1 day between consecutive populated days
    gaps = []
    import datetime as _dt
    for a, b in zip(days_sorted, days_sorted[1:]):
        try:
            da = _dt.date.fromisoformat(a); db = _dt.date.fromisoformat(b)
            delta = (db - da).days
            if delta > 1:
                gaps.append({"after": a, "before": b, "missing_days": delta - 1})
        except Exception:
            continue
    return {
        "total_rows": len(_csat_rows),
        "date_min": days_sorted[0] if days_sorted else None,
        "date_max": days_sorted[-1] if days_sorted else None,
        "rows_per_month": months,
        "gaps_over_1_day": gaps[:50],
        "distinct_days": len(days_sorted),
    }

@app.get("/health")
@app.head("/health")  # UptimeRobot sends HEAD requests — must support both
async def health():
    b, bf = cache_get("batch:all")
    i, fi = cache_get("intel:all")
    intel = i or {}
    lock  = get_lock()
    return {
        "status":        "ok",
        "api_key_set":   bool(API_KEY),
        "batch_cached":  b is not None,
        "batch_fresh":   bf,
        "intel_cached":  i is not None,
        "intel_fresh":   fi,
        "refresh_running": lock.locked(),
        "intel_data": {
            "articles":  len(intel.get("articles",{}).get("articles",[])),
            "videos":    len(intel.get("videos",{}).get("videos",[])),
            "categories":len(intel.get("categories",{}).get("categories",[])),
            "queries":   len(intel.get("search",{}).get("top_queries",[])),
        },
        "call_budget":   "15 fixed + up to 12 timeline calls per refresh (27 max), 1s gap each, every 2 hours",
        "admin_auth": {"required": ADMIN_TOKEN_REQUIRED, "configured": bool(ADMIN_TOKEN)},
        "version":       DASHBOARD_VERSION,
        "ts":            _utcnow().isoformat(),
    }

@app.get("/api/refresh")
async def api_refresh(admin_ok: bool = Depends(require_admin_token)):
    """Manually trigger a full data refresh.
    Safer default than /cache/clear because it never drops cached data.
    Safe to call any time — protected by lock so only one runs at a time.
    """
    if _is_test_mode():
        return {"status": "refresh skipped in test mode",
                "note": "Network refresh jobs are not spawned during automated tests."}
    lock = get_lock()
    if not lock.locked():
        asyncio.create_task(full_refresh())
        return {"status": "refresh started", "note": "Check /health for refresh status."}
    return {"status": "refresh already running", "note": "Check /health for progress."}

@app.get("/api/refresh/csat")
async def api_refresh_csat(x_admin_token: str = Header(default=""), authorization: str = Header(default="")):
    """Manually pull the latest CSAT data from Domo and rebuild the index.

    CSAT refresh is not locked by default so the UI can refresh Domo without an
    admin modal. Set CSAT_REFRESH_REQUIRES_ADMIN=true to restore the guard.
    """
    if CSAT_REFRESH_REQUIRES_ADMIN:
        await require_admin_token(x_admin_token=x_admin_token, authorization=authorization)
    if not _domo_configured():
        return JSONResponse(
            {"status": "domo not configured",
             "note": "Set DOMO_CLIENT_ID, DOMO_CLIENT_SECRET, DOMO_DATASET_ID env vars."},
            status_code=400,
        )
    ok = await asyncio.to_thread(_refresh_csat_from_domo)
    if ok:
        idx = _csat_index.get("index_data", {})
        return {"status": "csat refreshed from domo",
                "date_min": idx.get("date_min"),
                "date_max": idx.get("date_max"),
                "rows": idx.get("total_rows"),
                "source": CSAT_REFRESH_SOURCE}
    return JSONResponse(
        {"status": "domo pull failed", "note": "Check server logs for the [domo] error line."},
        status_code=502,
    )

@app.get("/cache/clear")
async def clear_cache(force: bool = False, admin_ok: bool = Depends(require_admin_token)):
    """Reset HTTP client and optionally clear cache before queueing refresh.

    Default behaviour keeps last-good cache to avoid blank dashboards during
    upstream outages. Use force=true only when you intentionally want to drop
    cached data immediately.
    """
    if _is_test_mode():
        if force:
            _cache.clear()
        return {"status": "cache test-mode no-op",
                "note": "Refresh jobs are not spawned during automated tests."}
    global _client
    if _client and not _client.is_closed:
        await _client.aclose()
    _client = None
    if force:
        _cache.clear()
    lock = get_lock()
    if not lock.locked():
        asyncio.create_task(full_refresh())
        if force:
            return {"status": "cache cleared + refresh queued", "note": "Cache dropped immediately; refresh takes ~30-60s."}
        return {"status": "cache retained + refresh queued", "note": "Safe refresh started without dropping last-good data."}
    if force:
        return {"status": "cache cleared — refresh already running", "note": "Dashboard may look empty until refresh succeeds."}
    return {"status": "cache retained — refresh already running", "note": "Last-good data preserved while refresh runs."}

@app.get("/debug/cms")
async def debug_cms(admin_ok: bool = Depends(require_admin_token)):
    client = get_client()
    try:
        r = await client.get(f"{CMS_BASE}/cs-portal-content-events/query/top-n",
                             params={"event":"article.viewed","groupBy":"itemId","n":3})
        cms_status = {"status": r.status_code, "preview": r.text[:150]}
    except Exception as e:
        cms_status = {"error": str(e)}
    b, _  = cache_get("batch:all")
    i, _  = cache_get("intel:all")
    intel = i or {}
    lock  = get_lock()
    return {
        "cms":           cms_status,
        "batch_cached":  b is not None,
        "intel_cached":  i is not None,
        "refresh_running": lock.locked(),
        "articles":      len(intel.get("articles",{}).get("articles",[])),
        "videos":        len(intel.get("videos",{}).get("videos",[])),
        "queries":       len(intel.get("search",{}).get("top_queries",[])),
    }

@app.get("/debug/video-topn")
async def debug_video_topn(admin_ok: bool = Depends(require_admin_token)):
    """Test whether the CMS supports grouping video.watched by properties.videoTitle.
    If it returns video titles, we can replace the timeline-sample approach with a
    single direct CMS call and get ALL videos, not just from sampled users.
    """
    client = get_client()
    results = {}
    # Test 1: group by properties.videoTitle (ideal — gives titles directly)
    try:
        r = await client.get(f"{CMS_BASE}/cs-portal-content-events/query/top-n",
                             params={"event":"video.watched","groupBy":"properties.videoTitle","n":10})
        results["by_videoTitle"] = {"status": r.status_code, "data": r.json() if r.status_code == 200 else r.text[:200]}
    except Exception as e:
        results["by_videoTitle"] = {"error": str(e)}
    await asyncio.sleep(1)
    # Test 2: group by itemId (known to be null — confirming)
    try:
        r = await client.get(f"{CMS_BASE}/cs-portal-content-events/query/top-n",
                             params={"event":"video.watched","groupBy":"itemId","n":10})
        results["by_itemId"] = {"status": r.status_code, "data": r.json() if r.status_code == 200 else r.text[:200]}
    except Exception as e:
        results["by_itemId"] = {"error": str(e)}
    await asyncio.sleep(1)
    # Test 3: group by userId (how many unique users watch videos)
    try:
        r = await client.get(f"{CMS_BASE}/cs-portal-content-events/query/top-n",
                             params={"event":"video.watched","groupBy":"userId","n":10})
        results["by_userId"] = {"status": r.status_code, "data": r.json() if r.status_code == 200 else r.text[:200]}
    except Exception as e:
        results["by_userId"] = {"error": str(e)}
    return results

# ─────────────────────────────────────────────────────────────────────────────
# Starter Guide Service — transparent proxy endpoints
# All calls forwarded to SG_BASE with no auth headers required.
# ─────────────────────────────────────────────────────────────────────────────

class _SgResponse:
    """Thin wrapper so callers can distinguish 404/empty from hard failures."""
    def __init__(self, status: int, body):
        self.status = status
        self.body   = body
    @property
    def ok(self):   return self.status == 200
    @property
    def not_found(self): return self.status == 404
    @property
    def empty(self):
        """True when the upstream returned 200 but the data list is empty."""
        if not self.ok:
            return False
        b = self.body
        if isinstance(b, dict):
            d = b.get("data")
            return isinstance(d, list) and len(d) == 0
        return isinstance(b, list) and len(b) == 0


async def _sg_get(path: str, params: dict = None, *, auth: bool = True) -> "_SgResponse":
    """Forward one GET request to the Starter Guide Service.
    Always returns an _SgResponse — never raises, never returns None.
    Callers inspect .ok / .not_found / .body instead of None-checking.

    auth=True (default): adds Authorization: Bearer <SG_API_TOKEN> when the
    token is configured — required for /api/v1/* endpoints.
    auth=False: skips auth header — used for /public/v1/* endpoints.
    """
    url = f"{SG_BASE}{path}"
    headers: dict = {}
    if auth and SG_API_TOKEN:
        headers["Authorization"] = f"Bearer {SG_API_TOKEN}"
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.get(url, params=params or {}, headers=headers)
            print(f"[sg] GET {path} -> {r.status_code}", flush=True)
            try:
                body = r.json()
            except Exception:
                body = {"raw": r.text[:500]}
            return _SgResponse(r.status_code, body)
    except Exception as ex:
        print(f"[sg] GET {path} connection error: {ex}", flush=True)
        return _SgResponse(503, {"error": str(ex)})


@app.get("/api/sg/journeys/active-customers")
async def sg_active_customers(limit: int = 20, skip: int = 0):
    """Proxy: GET /api/v1/journeys/active-customers.

    Returns every customer that has at least one active (in-progress) journey,
    each with its full journey objects embedded (title, locale, expectedGuides,
    completedIds, expiredIds, progress, startedOn). This is the entry point the
    Starter Guides tab uses to discover all active guides without needing a
    customer GID up front.

    NOTE: this route MUST stay declared before /api/sg/journeys/{customer_gid},
    otherwise FastAPI matches "active-customers" as a customer_gid path param.
    """
    resp = await _sg_get("/api/v1/journeys/active-customers",
                         {"limit": limit, "skip": skip})
    if resp.ok:
        return resp.body
    if resp.not_found:
        return {"data": [], "meta": {"total": 0}}
    raise HTTPException(status_code=resp.status,
                        detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")


@app.get("/api/sg/journeys/{customer_gid}")
async def sg_journeys(
    customer_gid: str,
    journeyId: str = "",
    from_date: str = "",
    to_date:   str = "",
    limit: int = 20,
    skip:  int = 0,
):
    """Proxy: GET /api/v1/journeys/:customerGid"""
    params: dict = {"limit": limit, "skip": skip}
    if journeyId:   params["journeyId"]  = journeyId
    if from_date:   params["from"]       = from_date
    if to_date:     params["to"]         = to_date
    resp = await _sg_get(f"/api/v1/journeys/{customer_gid}", params)
    if resp.ok or resp.not_found:
        # 404 → treat as empty result set (customer has no journeys yet)
        if resp.not_found:
            return {"data": [], "meta": {"total": 0}}
        return resp.body
    raise HTTPException(status_code=resp.status,
                        detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")


@app.get("/api/sg/journeys/{journey_id}/answers")
async def sg_journey_answers(journey_id: str):
    """Proxy: GET /api/v1/journeys/:journeyId/answers"""
    resp = await _sg_get(f"/api/v1/journeys/{journey_id}/answers")
    if resp.ok:
        return resp.body
    if resp.not_found:
        return {"journeyId": journey_id, "answers": {}, "guideCount": 0}
    raise HTTPException(status_code=resp.status,
                        detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")


@app.get("/api/sg/journeys/{journey_id}/starter-guides")
async def sg_journey_guides(journey_id: str, limit: int = 50, skip: int = 0):
    """Proxy: GET /api/v1/journeys/:journeyId/starter-guides"""
    resp = await _sg_get(f"/api/v1/journeys/{journey_id}/starter-guides",
                         {"limit": limit, "skip": skip})
    if resp.ok:
        return resp.body
    if resp.not_found:
        return {"data": [], "meta": {"total": 0}}
    raise HTTPException(status_code=resp.status,
                        detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")


@app.get("/api/sg/starter-guides/{guide_id}")
async def sg_guide_detail(guide_id: str):
    """Proxy: GET /api/v1/starter-guides/:id (authenticated, full detail with answers).
    Falls back to the public endpoint /public/v1/starter-guides/:id if no token is set.
    The public endpoint returns lockedSlides + progress + journeyId but omits answers.
    """
    if SG_API_TOKEN:
        resp = await _sg_get(f"/api/v1/starter-guides/{guide_id}", auth=True)
        if resp.ok:
            return resp.body
        # If auth fails (401/403), fall through to the public endpoint so the
        # dashboard still shows slide/progress data.
        if resp.status not in (401, 403):
            raise HTTPException(status_code=resp.status,
                                detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")

    # No token or auth failed — use unauthenticated public endpoint
    resp = await _sg_get(f"/public/v1/starter-guides/{guide_id}", auth=False)
    if resp.ok:
        return resp.body
    raise HTTPException(status_code=resp.status,
                        detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")


@app.get("/api/sg/public/starter-guides/{guide_id}")
async def sg_guide_detail_public(guide_id: str):
    """Proxy: GET /public/v1/starter-guides/:id — no auth required.
    Returns id, customerGid, journeyId, templateId, lockedSlides, progress,
    createdAt, expiresAt. Use when no SG_API_TOKEN is available.
    """
    resp = await _sg_get(f"/public/v1/starter-guides/{guide_id}", auth=False)
    if resp.ok:
        return resp.body
    raise HTTPException(status_code=resp.status,
                        detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")


@app.get("/api/sg/public/journeys/{journey_id}")
async def sg_journey_public(journey_id: str):
    """Proxy: GET /public/v1/journeys/:journeyId — no auth required.
    Returns journey overview for the Start Screen: title, subtitle, customerName,
    locale, footerText, helpUrl, expectedGuides, completedIds, expiredIds, progress.
    """
    resp = await _sg_get(f"/public/v1/journeys/{journey_id}", auth=False)
    if resp.ok:
        return resp.body
    raise HTTPException(status_code=resp.status,
                        detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")


@app.get("/api/sg/public/journeys/{journey_id}/starter-guides")
async def sg_journey_guides_public(journey_id: str):
    """Proxy: GET /public/v1/journeys/:journeyId/starter-guides — no auth required.
    Returns each expected guide with instanceId and status (PENDING/ACTIVE/COMPLETED/EXPIRED).
    """
    resp = await _sg_get(f"/public/v1/journeys/{journey_id}/starter-guides", auth=False)
    if resp.ok:
        return resp.body
    raise HTTPException(status_code=resp.status,
                        detail=f"Starter Guide Service error {resp.status}: {str(resp.body)[:200]}")


# ── CMS metrics for Starter Guides ───────────────────────────────────────────
# SG_CMS_PROJECT: set STARTER_GUIDE_CMS_PROJECT env var to match whatever project
# name your CMS uses for starter guide events. The Starter Guide Service emits its
# events under the "starter-guide-events" project (NOT the cs-portal-* namespace
# used by the rest of this dashboard).
SG_PROJECT = os.environ.get("STARTER_GUIDE_CMS_PROJECT", "starter-guide-events")


@app.get("/api/sg/metrics/timeseries")
async def sg_metrics_timeseries(event: str = "starter_guide.opened", bucket: str = "day"):
    """Time-series from CMS metrics for starter guide events."""
    series = await _timeseries(SG_PROJECT, event)
    return {"event": event, "project": SG_PROJECT, "series": series}


@app.get("/api/sg/metrics/topn")
async def sg_metrics_topn(event: str = "starter_guide.slide_viewed",
                           group_by: str = "itemId", n: int = 20):
    """Top-N from CMS metrics for starter guide events."""
    rows = await _topn(SG_PROJECT, event, group_by=group_by, n=n)
    return {"event": event, "project": SG_PROJECT, "groupBy": group_by, "top": rows}


@app.get("/api/sg/metrics/summary")
async def sg_metrics_summary(max_pages: int = 200, from_date: str = "", to_date: str = ""):
    """API-derived Starter Guide metrics — computed from the journeys service
    itself rather than from CMS event tracking.

    Pages through /api/v1/journeys/active-customers and aggregates the embedded
    journey data into the figures the Metrics dashboard needs:
      - activeCustomers / activeJourneys counts
      - guide totals + overall completion rate
      - funnel: how many journeys have 0,1,2,3,4+ guides completed
      - perGuide: completed vs expected count per guide template (with titles)
      - startsByDay: journeys started per calendar day (from startedOn)

    from_date / to_date (YYYY-MM-DD, inclusive) restrict the aggregate to the
    cohort of journeys whose trial started in that window (by startedOn), so the
    funnel/engagement reflect a chosen date range. Customer count after filtering
    is the number of distinct customers with a matching journey.

    This complements the CMS event charts (opens / answers) which can't be
    derived from the REST data.
    """
    PAGE = 100
    customers: list = []
    reachable = True
    # Fetch the first page to learn the total, then pull the rest concurrently so
    # the whole active-customer set is aggregated quickly (not one slow page at a time).
    first = await _sg_get("/api/v1/journeys/active-customers", {"limit": PAGE, "skip": 0})
    if first.ok:
        body = first.body or {}
        customers.extend(body.get("data", []) or [])
        total = int(((body.get("meta") or {}).get("total")) or len(customers))
        pages_needed = min(max(1, max_pages), -(-total // PAGE))  # ceil(total/PAGE), capped
        if pages_needed > 1:
            import asyncio
            results = await asyncio.gather(*[
                _sg_get("/api/v1/journeys/active-customers", {"limit": PAGE, "skip": p * PAGE})
                for p in range(1, pages_needed)
            ])
            for r in results:
                if r.ok:
                    customers.extend((r.body or {}).get("data", []) or [])
    else:
        reachable = first.not_found  # 404 = simply no active customers

    # Flatten every active journey across customers.
    journeys = [j for c in customers for j in (c.get("journeys") or [])]

    # Optional cohort filter by trial start date (startedOn), inclusive.
    if from_date or to_date:
        lo = (from_date or "")[:10]
        hi = (to_date or "")[:10]
        journeys = [j for j in journeys
                    if (d := (j.get("startedOn") or "")[:10])
                    and (not lo or d >= lo) and (not hi or d <= hi)]
    # Distinct customers represented after filtering.
    active_customers = len({j.get("customerGid") for j in journeys}) if (from_date or to_date) else len(customers)

    funnel = {"0": 0, "1": 0, "2": 0, "3": 0, "4+": 0}
    per_guide: dict = {}        # templateId -> {title, expected, completed}
    starts_by_day: dict = {}
    total_guides = completed_guides = 0

    for j in journeys:
        prog = j.get("progress") or {}
        total = int(prog.get("total") or 0)
        done  = int(prog.get("completed") or 0)
        total_guides     += total
        completed_guides += done

        funnel["4+" if done >= 4 else str(done)] = \
            funnel.get("4+" if done >= 4 else str(done), 0) + 1

        completed_ids = set(j.get("completedIds") or [])
        for eg in (j.get("expectedGuides") or []):
            tid = eg.get("templateId")
            if not tid:
                continue
            slot = per_guide.setdefault(tid, {
                "templateId": tid, "title": eg.get("title") or tid,
                "expected": 0, "completed": 0,
            })
            slot["expected"] += 1
            if tid in completed_ids:
                slot["completed"] += 1

        started = (j.get("startedOn") or "")[:10]
        if started:
            starts_by_day[started] = starts_by_day.get(started, 0) + 1

    completion_rate = round(completed_guides / total_guides * 100, 1) if total_guides else 0.0
    per_guide_list = sorted(per_guide.values(),
                            key=lambda r: r["completed"], reverse=True)
    starts_series = [{"date": d, "count": starts_by_day[d]}
                     for d in sorted(starts_by_day)]

    # ── Sequential step funnel + drop-off ────────────────────────────────
    # The 7 check-ins are delivered in order (Day 1, 4, 6, 14, 18, 24, 28) and a
    # later guide is only created after the previous is completed, so a journey
    # with progress.completed == N has cleared the first N check-ins. That lets us
    # build a true sequential funnel and pinpoint where customers stop.
    completed_counts = [int((j.get("progress") or {}).get("completed") or 0) for j in journeys]
    total_counts     = [int((j.get("progress") or {}).get("total") or 0) for j in journeys]
    max_steps = max(total_counts, default=0)
    # Canonical ordered step titles from the journey with the most expected guides.
    canonical, best_len = [], -1
    for j in journeys:
        egs = j.get("expectedGuides") or []
        if len(egs) > best_len:
            best_len = len(egs)
            canonical = [eg.get("title") or f"Step {i + 1}" for i, eg in enumerate(egs)]

    n_journeys = len(journeys)
    step_funnel, drop_off = [], []
    prev_completed = n_journeys
    for i in range(1, max_steps + 1):
        reached_i   = sum(1 for t in total_counts if t >= i)        # programme includes step i
        completed_i = sum(1 for c in completed_counts if c >= i)    # finished step i
        title = canonical[i - 1] if i - 1 < len(canonical) else f"Step {i}"
        step_funnel.append({"step": i, "title": title,
                            "reached": reached_i, "completed": completed_i})
        lost = max(prev_completed - completed_i, 0)
        drop_off.append({"step": i, "title": title, "lost": lost,
                         "lostPct": round(lost / prev_completed * 100, 1) if prev_completed else 0.0})
        prev_completed = completed_i
    biggest_drop = max(drop_off, key=lambda d: d["lost"]) if drop_off else None

    # ── Engagement breakdown (mutually exclusive across active journeys) ──
    not_started = in_progress = completed_journey = at_risk = 0
    for j in journeys:
        prog = j.get("progress") or {}
        total = int(prog.get("total") or 0)
        done  = int(prog.get("completed") or 0)
        if j.get("expiredIds"):
            at_risk += 1
        if done == 0:
            not_started += 1
        elif total and done >= total:
            completed_journey += 1
        else:
            in_progress += 1

    return {
        "source": "starter-guide-service-api",
        "reachable": reachable,
        "filter": {"from": from_date or None, "to": to_date or None},
        "activeCustomers": active_customers,
        "activeJourneys":  n_journeys,
        "totalGuides":     total_guides,
        "completedGuides": completed_guides,
        "completionRate":  completion_rate,
        "funnel":          funnel,
        "perGuide":        per_guide_list,
        "startsByDay":     starts_series,
        # New BA-oriented fields:
        "stepFunnel":      step_funnel,     # sequential reach/complete per check-in
        "dropOff":         drop_off,        # how many stop at each step
        "biggestDrop":     biggest_drop,    # the single worst drop-off point
        "engagement": {                     # mutually-exclusive status split
            "notStarted":  not_started,     # journey live, 0 check-ins completed
            "inProgress":  in_progress,     # 1..N-1 check-ins completed
            "completed":   completed_journey,
            "atRisk":      at_risk,          # has at least one expired guide
        },
        "stepTitles":      canonical,
    }


@app.get("/api/sg/metrics/delivery")
async def sg_metrics_delivery():
    """SMS delivery funnel for the Starter Guide.

    PLACEHOLDER until the messaging-provider endpoint is available. The Starter
    Guide service API does not expose SMS delivery receipts, so this returns a
    stable shape with `available: false` and null counts. The dashboard renders
    empty, clearly-labelled charts from it today.

    TO WIRE REAL DATA LATER: fetch the provider's delivery stats here, set
    `available = True`, fill `source` and the numeric fields below, and keep this
    exact response shape — the frontend will light up with no further changes.

        funnel.sent / received / notReceived / optedOut      -> top-of-funnel counts
        receivedBreakdown.started / completed / unfinished /
            droppedOut                                       -> what recipients did
        sentByDay: [{date, count}]                           -> optional time series
    """
    return {
        "available": False,
        "source": None,
        "funnel": {
            "sent":        None,
            "received":    None,
            "notReceived": None,
            "optedOut":    None,
        },
        "receivedBreakdown": {
            "started":    None,
            "completed":  None,
            "unfinished": None,
            "droppedOut": None,
        },
        "sentByDay": [],
    }


@app.get("/debug/sg")
async def debug_sg(admin_ok: bool = Depends(require_admin_token)):
    """One-shot health check for all Starter Guide layers.
    Hit this URL to diagnose why the dashboard shows no data.
    """
    result: dict = {
        "sg_base_url": SG_BASE,
        "sg_cms_project": SG_PROJECT,
        "auth": {
            "token_configured": bool(SG_API_TOKEN),
            "hint": (
                "Bearer token is set — authenticated /api/v1/* endpoints will be called."
                if SG_API_TOKEN
                else "No STARTER_GUIDE_API_TOKEN set. Only public /public/v1/* endpoints are usable. "
                     "Set STARTER_GUIDE_API_TOKEN in your Render env vars to enable full journey/answer data."
            ),
        },
    }

    # 1. Upstream reachability — use a known-safe path; 404 is fine, timeout is not
    probe = await _sg_get("/api/v1/journeys/debug-probe", {"limit": 1})
    result["upstream"] = {
        "reachable": probe.status not in (503,),
        "status":    probe.status,
        "body_preview": str(probe.body)[:200],
        "hint": (
            "Service is up (404 = customer not found, which is expected for a probe)"
            if probe.status in (200, 404)
            else "401/403 means the service is reachable but the Bearer token is missing or wrong."
            if probe.status in (401, 403)
            else "Service may be down or unreachable — check SG_BASE_URL and network egress"
        ),
    }

    # 2. CMS metrics — try to pull the last 1 data point for the SG project
    client = get_client()
    try:
        r = await client.get(
            f"{CMS_BASE}/{SG_PROJECT}/query/time-series",
            params={"event": "starter_guide.opened", "bucket": "day"},
        )
        result["cms_sg_project"] = {
            "status": r.status_code,
            "body_preview": r.text[:200],
            "hint": (
                "CMS project found and returning data"
                if r.status_code == 200
                else (
                    "CMS project not found — set STARTER_GUIDE_CMS_PROJECT env var "
                    f"to the correct project name (current: '{SG_PROJECT}'). "
                    "Known-good projects: cs-portal-content-events, cs-portal-auth-events, "
                    "cs-portal-feedback-events, cs-portal-items-events, cs-portal-scheduling-events. "
                    "Run /debug/sg/discover to auto-scan all projects for starter-guide events."
                )
            ),
        }
    except Exception as e:
        result["cms_sg_project"] = {"error": str(e)}

    return result


# Known CMS candidate event names that could contain starter guide tracking
_SG_CANDIDATE_EVENTS = [
    "starter_guide.opened",
    "starter_guide.viewed",
    "starter_guide.started",
    "starter_guide.completed",
    "starter_guide.slide_viewed",
    "starter_guide.answer_submitted",
    "guide.opened",
    "guide.viewed",
    "guide.started",
    "guide.completed",
    "guide.slide_viewed",
    "guide.answer_submitted",
    "onboarding.started",
    "onboarding.completed",
    "onboarding.guide_opened",
]

_KNOWN_CMS_PROJECTS = [
    "cs-portal-content-events",
    "cs-portal-auth-events",
    "cs-portal-feedback-events",
    "cs-portal-items-events",
    "cs-portal-scheduling-events",
]


@app.get("/debug/sg/discover")
async def debug_sg_discover(admin_ok: bool = Depends(require_admin_token)):
    """Scan every known CMS project for starter-guide events concurrently.
    Returns hits sorted by event count descending.
    Call this once to find the correct STARTER_GUIDE_CMS_PROJECT value.
    """
    client = get_client()

    async def probe(project: str, event: str):
        try:
            r = await client.get(
                f"{CMS_BASE}/{project}/query/time-series",
                params={"event": event, "bucket": "day"},
            )
            if r.status_code == 200:
                body   = r.json()
                series = body.get("series", [])
                total  = sum(int(p.get("count", 0) or 0) for p in series)
                if total > 0:
                    date_range = (
                        f"{series[0]['date']} → {series[-1]['date']}" if series else ""
                    )
                    print(f"[discover] HIT {project}/{event} = {total}", flush=True)
                    return {"project": project, "event": event,
                            "total_events": total, "data_points": len(series),
                            "date_range": date_range}
        except Exception as ex:
            print(f"[discover] {project}/{event} error: {ex}", flush=True)
        return None

    tasks = [
        probe(project, event)
        for project in _KNOWN_CMS_PROJECTS
        for event in _SG_CANDIDATE_EVENTS
    ]
    results = await asyncio.gather(*tasks)
    hits = [r for r in results if r is not None]
    hits.sort(key=lambda h: h["total_events"], reverse=True)

    if hits:
        best = hits[0]
        recommendation = (
            f"Set STARTER_GUIDE_CMS_PROJECT={best['project']} in your Render env vars "
            f"(found {best['total_events']} '{best['event']}' events there). "
            f"Then redeploy — no code change needed."
        )
    else:
        recommendation = (
            "No starter-guide events found in any known CMS project. "
            "Either events have not been tracked yet, or the project uses a name not in the known list. "
            "Run /debug/sg/projects to see all accessible CMS projects."
        )

    return {
        "probed_combinations": len(tasks),
        "hits_found": len(hits),
        "hits": hits,
        "recommendation": recommendation,
        "current_sg_project": SG_PROJECT,
        "known_projects_scanned": _KNOWN_CMS_PROJECTS,
        "events_probed": _SG_CANDIDATE_EVENTS,
    }


@app.get("/debug/sg/projects")
async def debug_sg_projects(admin_ok: bool = Depends(require_admin_token)):
    """List all CMS projects accessible with the current API key.
    Use this if /debug/sg/discover finds no hits — starter guide events
    may live in a project not in the known list.
    """
    client = get_client()
    results = {}

    async def check_project(project: str):
        try:
            r = await client.get(
                f"{CMS_BASE}/{project}/query/top-n",
                params={"event": "article.viewed", "groupBy": "itemId", "n": 1},
            )
            return project, {"status": r.status_code, "accessible": r.status_code == 200}
        except Exception as e:
            return project, {"status": "error", "accessible": False, "error": str(e)}

    project_results = await asyncio.gather(*[check_project(p) for p in _KNOWN_CMS_PROJECTS])
    for p, info in project_results:
        results[p] = info

    # Attempt a project-list endpoint if the CMS exposes one
    cms_list = None
    for list_path in ["/api/metrics/projects", "/api/projects"]:
        try:
            base = CMS_BASE.replace("/api/metrics", "")
            r = await client.get(f"{base}{list_path}")
            if r.status_code == 200:
                cms_list = r.json()
                break
        except Exception:
            pass

    return {
        "cms_base": CMS_BASE,
        "known_projects": results,
        "cms_project_list_endpoint": cms_list,
        "hint": (
            "Projects with accessible=true are queryable. "
            "If starter-guide events use a different project slug, "
            "set STARTER_GUIDE_CMS_PROJECT=<that-slug> in Render env vars."
        ),
    }


# ── Starter Guides page ───────────────────────────────────────────────────────

@app.get("/starter-guides", response_class=HTMLResponse)
async def starter_guides_page():
    """Starter Guides tab — served from starter_guides.html."""
    with open("starter_guides.html", encoding="utf-8") as f: html = f.read()
    return HTMLResponse(content=html,
        headers={"Cache-Control": "no-store, no-cache, must-revalidate",
                 "Pragma": "no-cache"})


@app.get("/", response_class=HTMLResponse)
async def dashboard():
    with open("index.html", encoding="utf-8") as f: html = f.read()
    return HTMLResponse(content=html,
        headers={"Cache-Control": "no-store, no-cache, must-revalidate",
                 "Pragma": "no-cache"})

@app.get("/csat", response_class=HTMLResponse)
async def csat_page():
    """Call Quality & CSAT page — served from csat.html."""
    with open("csat.html", encoding="utf-8") as f: html = f.read()
    return HTMLResponse(content=html,
        headers={"Cache-Control": "no-store, no-cache, must-revalidate",
                 "Pragma": "no-cache"})



def _dashboard_asset_response(filename: str, media_type: str):
    """Serve dashboard UI assets from root first, with /assets fallback for older deploys."""
    base = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(base, filename),
        os.path.join(base, "assets", filename),
    ]
    # Compatibility aliases: older HTML may still request /assets/dashboard_ux.*.
    if filename == "dashboard_ux.css":
        candidates.insert(0, os.path.join(base, "portal-overrides.css"))
    elif filename == "dashboard_ux.js":
        candidates.insert(0, os.path.join(base, "portal-system.js"))
    for asset_path in candidates:
        if os.path.exists(asset_path):
            with open(asset_path, encoding="utf-8") as f:
                content = f.read()
            return Response(content=content, media_type=media_type,
                headers={"Cache-Control": "no-store, no-cache, must-revalidate",
                         "Pragma": "no-cache"})
    raise HTTPException(status_code=404, detail="Asset not found")


@app.get("/portal-overrides.css")
async def portal_overrides_css():
    return _dashboard_asset_response("portal-overrides.css", "text/css")


@app.get("/portal-system.js")
async def portal_system_js():
    return _dashboard_asset_response("portal-system.js", "application/javascript")


@app.get("/assets/{asset_name}")
async def dashboard_asset(asset_name: str):
    """Serve shared dashboard UX assets with a whitelist and root-level fallback."""
    allowed = {
        "dashboard_ux.css": "text/css",
        "dashboard_ux.js": "application/javascript",
    }
    media_type = allowed.get(asset_name)
    if not media_type:
        raise HTTPException(status_code=404, detail="Asset not found")
    return _dashboard_asset_response(asset_name, media_type)

@app.get("/sw.js")
async def service_worker():
    with open("sw.js") as f: content = f.read()
    return Response(content=content, media_type="application/javascript",
        headers={"Cache-Control":"no-store, no-cache, must-revalidate",
                 "Pragma":"no-cache"})
